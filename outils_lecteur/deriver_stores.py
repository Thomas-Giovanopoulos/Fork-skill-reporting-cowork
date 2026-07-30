#!/usr/bin/env python3
"""Dérive les forme-stores de vérité (format convergé 2.1-skill) depuis les fixtures Excel.

Étape ② du chantier L (CDC v5 §4). Deux principes, chacun issu d'une décision :

**D47 — on réemploie l'`excel_to_store` de validation_app, on ne le réécrit pas.** Le module est
importé depuis le dépôt monté en LECTURE SEULE et exécuté tel quel ; sa sortie est la BASE du store.
Ce script ne fait que la **compléter et la converger** : les onglets que validation_app ne lit pas
(`Lignes`, `NC Flux` — les seuls présents dans les fixtures), les colonnes que son parseur perd
(type de poche, profil, mode, SRI, valeurs par poche, démembrement, date de souscription, montant
initial, uncalled), et les renommages du format convergé (`docs/confrontation_stores_2026-07-29.md`).

**Les positions ne se font confiance qu'en-têtes vérifiés.** La leçon du run Gronier (spec §7.4) :
un classeur à la topologie réduite est absorbé silencieusement par la projection d'en-têtes, et le
rendu s'affiche amputé sans que rien ne le signale. Ici c'est l'inverse : toute lecture positionnelle
est précédée d'un contrôle d'en-têtes (préfixe, insensible aux retours à la ligne), et un écart est
une **erreur fatale**, jamais une tolérance.

Usage :
    python3 deriver_stores.py [--fixtures DIR] [--sortie DIR] [--validation-app DIR]
                              [--quarter T2] [--year 2026]

Par défaut : les 7 fixtures du skill → `skill/p1_engine/tests/fixtures/stores/{stem}.store.json`,
période T2 2026 (celle de la régression). Chaque store est validé contre
`store_client.schema.json` (échec bruyant, L5) et ses références croisées vérifiées.
"""
from __future__ import annotations

import argparse
import importlib.util
import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

FORK = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FORK / "skill" / "pipeline"))
import store_builder  # noqa: E402

DEFAUT_VALIDATION_APP = "/sessions/epic-happy-ritchie/mnt/validation_app"

# ---------------------------------------------------------------------------
# En-têtes attendus (ligne 3), par préfixe d'onglet — préfixe suffisant, les
# libellés portent des dates variables (« Valeur au [Date] ») et des retours à
# la ligne. Vérifiés AVANT toute lecture positionnelle.
# ---------------------------------------------------------------------------

ENTETES = {
    "Fin coté": ["Nature", "Assureur", "Intermédiaire", "Type poche", "Société de gestion",
                 "Dépositaire", "Mode", "Profil", "Date d'invest", "Nantissement",
                 "Nominal", "Valeur 01/01", "Valeur au", "Classe", "Géographie"],
    "Lignes": ["Contrat", "Libellé", "ISIN", "Valeur", "Perf", "Poche", "Classe",
               "Géographie", "SRI"],
    "NC Flux": ["Nom du fonds", "Date", "Type", "Montant"],
    "Liq": ["Intitulé", "Banque", "Solde"],
    "Immo": ["Nom du bien", "Fonction", "Propriété", "Hypothèque", "Loyer",
             "Date acquisition", "Valeur acquisition", "Valeur au"],
    "Dettes": ["Intitulé", "Établissement", "Type", "Date souscription", "Montant initial",
               "Taux", "Échéance", "Périodicité", "Garantie", "Capital restant", "Adossement"],
    "Non coté": ["Nom du fonds", "Gestionnaire", "Stratégie", "Cible MOIC", "Capital engagé",
                 "Capital appelé", "Non appelé réel", "Non appelé estimé", "MOIC réalisé",
                 "Valeur", "Classe"],
}

TYPES_FLUX_NC = {
    "appel": "appel", "distribution": "distribution",
    "appel prevu": "appel_prevu", "appel prévu": "appel_prevu",
    "distribution prevue": "distribution_prevue", "distribution prévue": "distribution_prevue",
}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split()).casefold()


def verifier_entetes(ws, prefixe: str) -> None:
    attendus = ENTETES.get(prefixe)
    if not attendus:
        return
    ligne = [(_norm(c.value) if c.value is not None else "") for c in ws[3][: len(attendus)]]
    for i, attendu in enumerate(attendus):
        if not ligne[i].startswith(_norm(attendu)):
            raise SystemExit(
                f"ERREUR en-têtes [{ws.title}] col {i}: attendu un libellé commençant par "
                f"{attendu!r}, trouvé {ws[3][i].value!r}. La lecture positionnelle serait "
                f"fausse en silence — on s'arrête (leçon du classeur Gronier réduit, spec §7.4).")


# ---------------------------------------------------------------------------
# Petites conversions — mêmes conventions que le format convergé
# ---------------------------------------------------------------------------

def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


def _f(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _bool_oui(v) -> bool:
    return _norm(_s(v)) in ("oui", "o", "yes", "true", "x")


def _iso(v) -> str | None:
    """Date ISO si convertible, None sinon (l'appelant décide du verbatim)."""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = _s(v)
    m = re.fullmatch(r"(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    return None


def _date_ou_verbatim(v) -> str | None:
    """ISO quand convertible, VERBATIM sinon — le moteur affiche tel quel (« Févr. 2024 »)."""
    iso = _iso(v)
    if iso:
        return iso
    s = _s(v)
    return s or None


def _moic_nombre(v):
    """« 2,0x » / « 1.8x » / 1.8 → nombre (A3 : nombre au store, le lecteur formate)."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _s(v).lower().replace(",", ".").rstrip("x").strip()
    try:
        return float(s)
    except ValueError:
        return None  # l'appelant signale — jamais de valeur inventée


def _classe_code(libelle: str, label_to_class: dict) -> str | None:
    """Code de classe si le libellé mappe, VERBATIM sinon, absent si vide.

    Le `LABEL_TO_CLASS.get(x, "actions")` de validation_app est un DÉFAUT SILENCIEUX :
    une cellule vide devenait « actions ». Ici : vide → absent (ABSENCE ≠ NULL).
    """
    if not libelle:
        return None
    return label_to_class.get(libelle, libelle)


def _rows(ws, start=4):
    for row in ws.iter_rows(min_row=start, values_only=True):
        if not row or all(v is None for v in row):
            break
        yield row


def _feuilles(wb, prefixe: str, suffixe: str):
    for sn in wb.sheetnames:
        parts = sn.split(" — ")
        if parts[0].strip() == prefixe and (len(parts) == 1 or parts[1].strip() == suffixe):
            return wb[sn]
    return None


# ---------------------------------------------------------------------------
# Chargement du module validation_app (D47 : exécuté, jamais copié)
# ---------------------------------------------------------------------------

def charger_excel_to_store(rep_validation_app: Path):
    chemin = rep_validation_app / "ingest" / "excel_to_store.py"
    if not chemin.exists():
        raise SystemExit(f"ERREUR: excel_to_store.py introuvable sous {rep_validation_app} "
                         f"(--validation-app ou $VALIDATION_APP_DIR)")
    spec = importlib.util.spec_from_file_location("va_excel_to_store", chemin)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ---------------------------------------------------------------------------
# Convergence d'un store émis par validation_app vers 2.1-skill
# ---------------------------------------------------------------------------

def converger(base: dict, wb, va_mod, avertit) -> dict:
    """Transforme la sortie d'excel_to_store en store 2.1-skill, en re-lisant le
    classeur pour les colonnes/onglets que validation_app perd."""
    import openpyxl  # noqa: F401 — wb déjà ouvert par l'appelant

    # --- entités : ids CONSERVÉS tels que l'onglet Entités les déclare -------
    # Le manifeste (excel_to_manifest) reprend ces mêmes ids : les re-frapper en
    # tmp_ent_NNN casserait la correspondance manifeste ↔ store au rendu (constat
    # du 30/07, étape ③ — le motif du schéma a été assoupli en conséquence).
    suffixe_de = {}
    remap = {}
    ws_ent = None
    for sn in wb.sheetnames:
        if "entit" in sn.lower():
            ws_ent = wb[sn]
            break
    lignes_ent = list(_rows(ws_ent)) if ws_ent is not None else []
    entites = []
    for i, ent in enumerate(base["client"]["entities"], start=1):
        remap[ent["id"]] = ent["id"]
        entites.append({"id": ent["id"], "label": ent["label"], "type": ent["type"],
                        "order": ent["order"]})
        suf = _s(lignes_ent[i - 1][3]) if i - 1 < len(lignes_ent) else "PP"
        suffixe_de[ent["id"]] = suf or "PP"

    reporting = {k: v for k, v in base["reporting"].items() if k != "blocs_enabled"}
    # Profil : validation_app le DEVINE depuis la colonne 7 du premier contrat — heuristique
    # divergente de celle du moteur (excel_to_manifest : cellule « Profil » de l'onglet Entités,
    # sinon Équilibré). Attrapé par la preuve d'équivalence des manifestes (bascule ⑤, 30/07).
    # On reproduit l'heuristique du MOTEUR : c'est elle que le golden protège.
    profil_cellule = None
    if ws_ent is not None:
        for row in ws_ent.iter_rows(values_only=True):
            for i, c in enumerate(row or []):
                if isinstance(c, str) and "profil" in c.lower():
                    if ":" in c:
                        profil_cellule = c.split(":", 1)[1].strip()
                    elif i + 1 < len(row) and row[i + 1]:
                        profil_cellule = str(row[i + 1]).strip()
                    break
            if profil_cellule:
                break
    reporting["profile"] = profil_cellule or "Équilibré"

    store = {
        "schema_version": "2.1-skill",
        "client": {"id": "tmp_c_001", "label": base["client"]["label"], "entities": entites},
        "reporting": reporting,
        "_provisional_ids": True,
    }

    # --- financier coté : base v-app enrichie par relecture du classeur ------
    fc_conv, valuations, mouvements = [], [], []
    pck_seq = [0]
    date_rep = reporting["date_reporting"]

    for ent in entites:
        eid, suf = ent["id"], suffixe_de[ent["id"]]
        ws = _feuilles(wb, "Fin coté", suf)
        if ws is None:
            continue
        verifier_entetes(ws, "Fin coté")
        contrats = {}
        for row in _rows(ws):
            nature = _s(row[0])
            if not nature:
                continue
            assureur, inter = _s(row[1]), _s(row[2])
            cle = (nature.casefold(), assureur, inter)
            if cle not in contrats:
                contrats[cle] = {
                    "id": f"tmp_fc_{len(fc_conv) + len(contrats) + 1:03d}",
                    "entity_id": eid,
                    "label": f"{nature} — {assureur}",
                    "nature": nature,  # VERBATIM — c'est le mot qui fait les clés de jointure (30/07)
                    "assureur": assureur,
                    "intermediaire": inter or None,
                    "envelope_type": va_mod._envelope_from_nature(nature, _s(row[5])),
                    "manager": _s(row[4]),
                    "custodian": _s(row[5]) or None,
                    "management_mode": _s(row[6]) or None,
                    "risk_profile": _s(row[7]) or None,
                    "invest_date": _date_ou_verbatim(row[8]),
                    "nantissement": _bool_oui(row[9]),
                    "source": "excel",
                    "attributes": {"pockets": []},
                    "_nature": nature,
                }
            c = contrats[cle]
            pck_seq[0] += 1
            n_cols = len(row)
            libelle_poche = _s(row[16]) if n_cols > 16 else ""
            poche = {
                "id": f"pck_{pck_seq[0]:03d}",
                "label": libelle_poche or c["label"],
                "type": _s(row[3]) or None,
                "profile": _s(row[7]) or None,
                "manager": _s(row[4]) or None,
                "custodian": _s(row[5]) or None,
                "value": _f(row[12]) if _f(row[12]) is not None else 0.0,
                "capital_invested": _f(row[10]),
                "value_jan1": _f(row[11]),
                "invest_date": _date_ou_verbatim(row[8]),
                "nantissement": _bool_oui(row[9]) or None,
                "classe_rhetores": _classe_code(_s(row[13]), va_mod.LABEL_TO_CLASS),
                "geography": (_s(row[14]) or None) if n_cols > 14 else None,
                "sri": (int(row[15]) if n_cols > 15 and isinstance(row[15], (int, float)) else None),
            }
            c["attributes"]["pockets"].append(poche)
            if _f(row[11]) is not None:
                valuations.append({"position_id": poche["id"],
                                   "date": f"{date_rep[:4]}-01-01", "value": _f(row[11])})
            valuations.append({"position_id": poche["id"], "date": date_rep,
                               "value": poche["value"]})

        # agrégats contrat + jointure Lignes
        ws_lignes = _feuilles(wb, "Lignes", suf)
        lignes_par_cle = {}
        if ws_lignes is not None:
            verifier_entetes(ws_lignes, "Lignes")
            for lr in _rows(ws_lignes):
                lignes_par_cle.setdefault(_s(lr[0]), []).append(lr)

        for c in contrats.values():
            poches = c["attributes"]["pockets"]
            c["value_current"] = round(sum(p["value"] for p in poches), 2)
            caps = [p["capital_invested"] for p in poches]
            if all(v is not None for v in caps):
                c["capital_invested"] = round(sum(caps), 2)
            v01 = [p["value_jan1"] for p in poches]
            if all(v is not None for v in v01):
                c["value_jan1"] = round(sum(v01), 2)

            cle_lignes = c.pop("_nature") and c["label"]
            if cle_lignes in lignes_par_cle:
                poche_par_label = {p["label"]: p["id"] for p in poches}
                lines = []
                for lr in lignes_par_cle.pop(cle_lignes):
                    lbl_poche = _s(lr[5])
                    if lbl_poche and lbl_poche not in poche_par_label:
                        raise SystemExit(
                            f"ERREUR Lignes [{suf}]: la poche {lbl_poche!r} de la ligne "
                            f"{_s(lr[1])!r} ne correspond à aucune poche du contrat "
                            f"{c['label']!r} — jointure par id impossible (C7/D16).")
                    classe = _classe_code(_s(lr[6]), va_mod.LABEL_TO_CLASS)
                    if classe is None:
                        raise SystemExit(
                            f"ERREUR Lignes [{suf}]: classe vide sur {_s(lr[1])!r} — "
                            f"`class` est requis sur une ligne, pas de défaut silencieux.")
                    lines.append({k: v for k, v in {
                        "isin": _s(lr[2]) or None,
                        "label": _s(lr[1]),
                        "value": _f(lr[3]) if _f(lr[3]) is not None else 0.0,
                        "class": classe,
                        "geography": _s(lr[7]) or None,
                        "sri": int(lr[8]) if isinstance(lr[8], (int, float)) else None,
                        "pocket": poche_par_label.get(lbl_poche) if lbl_poche else None,
                        "perf_pct": _f(lr[4]),
                    }.items() if v is not None})
                c["attributes"]["lines"] = lines
            fc_conv.append({k: v for k, v in c.items() if v is not None})

        if ws_lignes is not None and lignes_par_cle:
            orphelines = list(lignes_par_cle)
            raise SystemExit(f"ERREUR Lignes [{suf}]: clés de contrat sans contrat: {orphelines}")

    store["financier_cote"] = fc_conv

    # --- catégories typées : base v-app + colonnes perdues --------------------
    idx_par_entite: dict[tuple, int] = {}

    def _patch(categorie, prefixe, patcheur):
        sortie = []
        for item in base.get(categorie, []):
            eid = remap[item["entity_id"]]
            suf = suffixe_de[eid]
            ws = _feuilles(wb, prefixe, suf)
            if ws is None:
                raise SystemExit(f"ERREUR: entrée {categorie} sans onglet {prefixe} — {suf}")
            verifier_entetes(ws, prefixe)
            k = (categorie, eid)
            idx = idx_par_entite.get(k, 0)
            idx_par_entite[k] = idx + 1
            row = list(_rows(ws))[idx]
            item = dict(item)
            item["entity_id"] = eid
            patcheur(item, row)
            sortie.append({k2: v for k2, v in item.items() if v is not None})
        return sortie

    def _patch_immo(item, row):
        item.pop("ownership_pct", None)  # codé en dur chez v-app : perte silencieuse corrigée
        item["ownership"] = _s(row[2]) or None
        item.setdefault("attributes", {})

    def _patch_dette(item, row):
        item["date_souscription"] = _date_ou_verbatim(row[3])
        item["montant_initial"] = _f(row[4])
        # Le taux du classeur peut être une FORMULE (« SOFR+0,9% ») : le nombre de
        # validation_app n'en est qu'une projection, le verbatim est la donnée.
        item["taux"] = _s(row[5]) or None
        if item.get("capital_remaining") is None:
            raise SystemExit(f"ERREUR Dettes: capital restant dû absent sur {item['label']!r} "
                             f"— il entre dans les contrôles comptables (MQ6).")

    def _patch_nc(item, row):
        attrs = dict(item.get("attributes") or {})
        for cle in ("moic_target", "moic_realise"):
            if cle in attrs:
                n = _moic_nombre(attrs[cle])
                if n is None:
                    avertit(f"{item['label']}: {cle} {attrs[cle]!r} non convertible en nombre "
                            f"(A3) — champ omis, à corriger à la source")
                    attrs.pop(cle)
                else:
                    attrs[cle] = n
        u_reel, u_est = _f(row[6]), _f(row[7])
        if u_reel is not None:
            attrs["uncalled_reel"] = u_reel
        if u_est is not None:
            attrs["uncalled_estime"] = u_est
        item["attributes"] = attrs or None

    def _patch_liq(item, row):
        # `_float(row[2]) or 0.0` chez validation_app : un solde VIDE devenait 0 €.
        # Or le moteur affiche « — » pour une cellule vide — ABSENCE ≠ NULL ≠ zéro.
        if row[2] is None or _s(row[2]) == "":
            item.pop("balance", None)

    store["liquidites"] = _patch("liquidites", "Liq", _patch_liq)
    store["immobilier"] = _patch("immobilier", "Immo", _patch_immo)
    store["dettes"] = _patch("dettes", "Dettes", _patch_dette)
    store["non_cote"] = _patch("non_cote", "Non coté", _patch_nc)

    # --- NC Flux → mouvements (MQ3) ------------------------------------------
    nc_par_label = {i["label"]: i["id"] for i in store["non_cote"]}
    for ent in entites:
        ws = _feuilles(wb, "NC Flux", suffixe_de[ent["id"]])
        if ws is None:
            continue
        verifier_entetes(ws, "NC Flux")
        for row in _rows(ws):
            nom = _s(row[0])
            if not nom:
                continue
            if nom not in nc_par_label:
                raise SystemExit(f"ERREUR NC Flux: {nom!r} ne correspond à aucune entrée non coté "
                                 f"(jointure EXACTE, spec §3) — entrées connues: {list(nc_par_label)}")
            montant = _f(row[3])
            if montant is None or montant < 0:
                raise SystemExit(f"ERREUR NC Flux: montant invalide {row[3]!r} sur {nom!r} "
                                 f"(A4 : toujours positif, le sens vient du type)")
            d = _iso(row[1])
            if d is None:
                raise SystemExit(f"ERREUR NC Flux: date invalide {row[1]!r} sur {nom!r}")
            type_brut = _norm(_s(row[2]))
            # « Valorisation » n'est PAS un flux : c'est un point de VL par fonds — son
            # logement convergé est valuations[] (position_id = l'entrée non coté).
            # Découvert par l'échec bruyant du 30/07 : MQ3 n'avait recensé que les flux,
            # FLUX_TYPES du moteur en connaît cinq.
            if type_brut == "valorisation":
                valuations.append({"position_id": nc_par_label[nom], "date": d,
                                   "value": montant})
                continue
            t = TYPES_FLUX_NC.get(type_brut)
            if t is None:
                raise SystemExit(f"ERREUR NC Flux: type {_s(row[2])!r} inconnu "
                                 f"(attendus: {sorted(set(TYPES_FLUX_NC.values()))} + Valorisation)")
            mouvements.append({"id": f"tmp_mv_{len(mouvements) + 1:03d}",
                               "entry_ref": nc_par_label[nom], "date": d,
                               "type": t, "amount": montant})
    if mouvements:
        store["mouvements"] = mouvements

    # --- séries : telles quelles depuis v-app, positions re-visées -----------
    if valuations:
        store["valuations"] = valuations
    if base.get("courbe_performance"):
        store["courbe_performance"] = base["courbe_performance"]
    if base.get("historique_annuel"):
        store["historique_annuel"] = base["historique_annuel"]

    # nettoyage ABSENCE ≠ NULL (mêmes règles que store_builder)
    return store_builder._clean(store)


# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--fixtures", default=str(FORK / "skill/p1_engine/tests/fixtures"))
    ap.add_argument("--sortie", default=str(FORK / "skill/p1_engine/tests/fixtures/stores"))
    ap.add_argument("--validation-app",
                    default=os.environ.get("VALIDATION_APP_DIR", DEFAUT_VALIDATION_APP))
    ap.add_argument("--quarter", default="T2")
    ap.add_argument("--year", type=int, default=2026)
    args = ap.parse_args()

    import openpyxl

    va_mod = charger_excel_to_store(Path(args.validation_app))
    sortie = Path(args.sortie)
    sortie.mkdir(parents=True, exist_ok=True)

    fixtures = sorted(Path(args.fixtures).glob("fx_*.xlsx"))
    if not fixtures:
        raise SystemExit(f"Aucune fixture sous {args.fixtures}")

    echecs = 0
    for fx in fixtures:
        avertissements: list[str] = []
        base = va_mod.excel_to_store(fx, args.quarter, args.year)
        wb = openpyxl.load_workbook(fx, data_only=True)
        store = converger(base, wb, va_mod, avertissements.append)

        erreurs = store_builder.validate_all(store)
        refs = store_builder.check_refs(store)
        if erreurs or refs:
            print(f"ÉCHEC  {fx.stem}")
            for e in erreurs + refs:
                print(f"       {e}")
            echecs += 1
            continue

        dest = sortie / f"{fx.stem}.store.json"
        dest.write_text(json.dumps(store, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        n_fc = len(store.get("financier_cote", []))
        n_pck = sum(len(c["attributes"]["pockets"]) for c in store.get("financier_cote", []))
        tot = sum(c["value_current"] for c in store.get("financier_cote", []))
        print(f"OK     {fx.stem}: {n_fc} contrat(s)/{n_pck} poche(s) = {tot:,.0f} € · "
              f"{len(store.get('non_cote', []))} nc · {len(store.get('liquidites', []))} liq · "
              f"{len(store.get('immobilier', []))} immo · {len(store.get('dettes', []))} dette(s) · "
              f"{len(store.get('mouvements', []))} mvt(s)")
        for a in avertissements:
            print(f"       ⚠ {a}")

    print(f"\n{'ÉCHEC' if echecs else 'OK'} — {len(fixtures) - echecs}/{len(fixtures)} store(s) dérivé(s)")
    return 1 if echecs else 0


if __name__ == "__main__":
    sys.exit(main())
