#!/usr/bin/env python3
"""Lint du fichier Excel source — contrôle de saisie AVANT génération.

Usage : python3 lint.py source.xlsx
Sortie : liste d'ERREURS (bloquantes) et d'AVERTISSEMENTS (non bloquants).
Code retour 1 si au moins une erreur.
"""
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8"); sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass
import sys, re
import openpyxl
try:
    import colmap as _cm
except ImportError:
    from pathlib import Path as _P; sys.path.insert(0, str(_P(__file__).parent)); import colmap as _cm

def _rows_proj(ws, kind):
    """Lignes de données projetées vers les indices canoniques (repli identité)."""
    try: mapping=_cm.build_mapping([c.value for c in ws[3]], kind)
    except Exception: mapping=None
    out=[]
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or not any(v not in (None,"") for v in r): continue
        a=str(r[0] or "").strip()
        if a.startswith(("•","ℹ","LÉGENDE","LEGENDE")) or a.upper().startswith("TOTAL"): continue
        out.append(_cm.project(list(r), mapping, kind) if mapping else list(r))
    return out

CATS = ["liquidites", "immobilier", "financier_cote", "non_cote", "dettes"]
PREFIX = {"Liq": "liquidites", "Immo": "immobilier", "Fin coté": "financier_cote",
          "Non coté": "non_cote", "Dettes": "dettes"}
CLASSES_12 = {"Actions","Obligations","Produits structurés","Fonds euros","Alternatifs",
              "Matières premières","Crypto","Monétaire","Private Equity","Dette privée",
              "Immo non coté","Immobilier non coté","Infrastructures"}
CLASSES_NC = {"PE","Private Equity","Dette privée","Immo non coté","Immobilier non coté","Infrastructures"}
GEOS = {"Amérique du Nord","Europe développée","International / Monde","Émergents","Asie-Pacifique"}

def isnum(v): return isinstance(v, (int, float))
def nonempty(v): return v not in (None, "")

def lint(path):
    iss = []  # (level, where, msg)
    E = lambda w, m: iss.append(("ERREUR", w, m))
    W = lambda w, m: iss.append(("AVERT.", w, m))
    wb = openpyxl.load_workbook(path, data_only=True)

    if "Entités" not in wb.sheetnames:
        E("Entités", "onglet 'Entités' manquant — impossible de déterminer les entités.")
        return iss
    ents = []
    for r in wb["Entités"].iter_rows(min_row=4, values_only=True):
        if not r or not r[0]: continue
        eid, label, etype, suf = (list(r) + [None]*4)[:4]
        ents.append({"id": str(eid).strip(), "label": str(label).strip() if label else "",
                     "type": str(etype).strip().lower() if etype else "", "suf": str(suf).strip() if suf else ""})
    if not ents:
        E("Entités", "aucune entité déclarée."); return iss
    pps = [e for e in ents if e["type"] == "pp"]
    if len(pps) > 1: E("Entités", f"au plus 1 PP attendue, trouvé {len(pps)}.")
    if pps and ents[0]["type"] != "pp": E("Entités", "la PP doit être en première ligne.")
    ids = [e["id"] for e in ents]
    if len(ids) != len(set(ids)): E("Entités", f"id d'entité dupliqués : {ids}.")
    for e in ents:
        if not re.match(r"^[a-z0-9_]+$", e["id"]): W("Entités", f"id '{e['id']}' non snake_case.")
        if e["type"] not in ("pp", "holding"): E("Entités", f"type invalide '{e['type']}' pour {e['id']}.")
        if not e["label"]: W("Entités", f"libellé vide pour {e['id']}.")
        if not e["suf"]: E("Entités", f"suffixe d'onglet vide pour {e['id']}.")

    # onglets catégorie
    sheets = set(wb.sheetnames)
    # schéma v2 : contrats dont les Lignes portent une Classe (col 7) -> classe Fin coté facultative
    NATURE_L = {"av": "Assurance-vie", "capi": "Contrat de capitalisation", "cto": "CTO"}
    _nrm = lambda x: " ".join(str(x or "").split()).casefold()
    classed = {}   # suf -> set(clé contrat normalisée)
    lines_cls = {} # suf -> list((ck, i, cls, geo, valeur_ok))
    for e in ents:
        lname = f"Lignes — {e['suf']}"
        cks = set(); lcs = []
        if lname in sheets:
            for i, r in enumerate(_rows_proj(wb[lname], "Lignes"), 1):
                gg = lambda j: r[j] if j < len(r) else None
                if not nonempty(gg(0)) or not nonempty(gg(1)): continue
                c7 = gg(6)
                if nonempty(c7):
                    cks.add(_nrm(gg(0)))
                    lcs.append((f"{lname} L{i+3}", str(c7).strip(), gg(7)))
        classed[e["suf"]] = cks; lines_cls[e["suf"]] = lcs
        for loc, c7, g8 in lcs:
            if c7 not in CLASSES_12 and c7 != "Monétaire": W(loc, f"classe inconnue '{c7}' (typo ?).")
            if nonempty(g8) and str(g8).strip() not in GEOS: W(loc, f"géographie inconnue '{g8}'.")
    for e in ents:
        present = [c for c in CATS if any(f"{p} — {e['suf']}" in sheets for p, cc in PREFIX.items() if cc == c)]
        if not present:
            W(e["label"] or e["id"], "aucun onglet catégorie avec données.")
        for prefix, cat in PREFIX.items():
            name = f"{prefix} — {e['suf']}"
            if name not in sheets: continue
            ws = wb[name]
            if prefix in ("Fin coté","Non coté"):
                rows = _rows_proj(ws, prefix)
            else:
                rows = [list(r) for r in ws.iter_rows(min_row=4, values_only=True) if r and any(nonempty(c) for c in r)]
            for i, r in enumerate(rows, 1):
                loc = f"{name} L{i+3}"
                g = lambda j: r[j] if j < len(r) else None
                if cat == "liquidites":
                    if not isnum(g(2)): E(loc, f"solde non numérique : {g(2)!r}.")
                    if not nonempty(g(1)): W(loc, "banque vide.")
                elif cat == "immobilier":
                    if not isnum(g(7)): E(loc, f"valeur non numérique : {g(7)!r}.")
                elif cat == "financier_cote":
                    if not isnum(g(10)): W(loc, f"nominal non numérique : {g(10)!r}.")
                    if not isnum(g(12)): E(loc, f"valeur non numérique : {g(12)!r}.")
                    if nonempty(g(11)) and not isnum(g(11)): W(loc, f"valeur 01/01 non numérique : {g(11)!r}.")
                    cls = g(13)
                    _nat = NATURE_L.get(str(g(0) or "").strip().lower(), str(g(0) or "").strip())
                    _isclassed = _nrm(f"{_nat} — {g(1)}") in classed.get(e["suf"], set())
                    if not nonempty(cls):
                        if not _isclassed: W(loc, "classe vide et aucune ligne classée pour ce contrat (donut classes incomplet).")
                    elif str(cls).strip() not in CLASSES_12: E(loc, f"classe inconnue '{cls}' (typo ?).")
                    if not _isclassed and str(cls).strip() == "Actions" and not nonempty(g(14)):
                        W(loc, "action sans géographie (donut géo incomplet).")
                    if nonempty(g(14)) and str(g(14)).strip() not in GEOS: W(loc, f"géographie inconnue '{g(14)}'.")
                    if not nonempty(g(5)): W(loc, "dépositaire vide.")
                elif cat == "non_cote":
                    if not isnum(g(4)): W(loc, f"capital engagé non numérique : {g(4)!r}.")
                    if str(g(10)).strip() not in CLASSES_NC: W(loc, f"classe non coté inattendue '{g(10)}'.")
                elif cat == "dettes":
                    if not isnum(g(9)) and not isnum(g(4)): E(loc, "capital restant et montant initial non numériques.")
                    ad = g(10)
                    if nonempty(ad) and str(ad).strip() not in CATS: W(loc, f"adossement inconnu '{ad}'.")

    # onglets Mouvements : jointure clé contrat, type, montant, date
    MVT_T = ("versement","apport","souscription","retrait","rachat","frais")
    for e in ents:
        mname = f"Mouvements — {e['suf']}"
        if mname not in sheets: continue
        fname = f"Fin coté — {e['suf']}"
        cles = set()
        NATL = {"av":"assurance-vie","capi":"contrat de capitalisation","cto":"cto"}
        if fname in sheets:
            for fr in _rows_proj(wb[fname], "Fin coté"):
                nat=str(fr[0] or "").strip(); ass=str(fr[1] or "").strip()
                if nat and ass: cles.add(_nrm(f"{NATL.get(nat.lower(),nat)} — {ass}"))
        for i, r in enumerate(_rows_proj(wb[mname], "Mouvements"), 1):
            loc=f"{mname} L{i+3}"
            gg=lambda j: r[j] if j < len(r) else None
            if cles and _nrm(gg(1)) not in cles: E(loc, f"contrat '{gg(1)}' absent du Fin coté (jointure impossible).")
            t=str(gg(2) or "").strip().casefold()
            if not any(t.startswith(x) for x in MVT_T): E(loc, f"type de mouvement inconnu '{gg(2)}' (Versement / Retrait / Frais).")
            if not isnum(gg(3)): E(loc, f"montant invalide : {gg(3)!r}.")

    # onglets NC Flux : jointure fonds, types, montants, dates
    FLUXT = {"Appel","Distribution","Valorisation","Appel prévu","Distribution prévue"}
    import datetime as _dt
    def _pd(v):
        if isinstance(v, _dt.datetime) or isinstance(v, _dt.date): return True
        try:
            d,m,y=(int(x) for x in str(v).strip()[:10].split("/")); return 1<=d<=31 and 1<=m<=12
        except Exception: return False
    for e in ents:
        name = f"NC Flux — {e['suf']}"
        if name not in sheets: continue
        ncname = f"Non coté — {e['suf']}"
        funds = set()
        if ncname in sheets:
            for r in wb[ncname].iter_rows(min_row=4, values_only=True):
                if r and nonempty(r[0]): funds.add(str(r[0]).strip())
        def _keep(r):
            if not r or not any(nonempty(c) for c in r): return False
            a=str(r[0] or "").strip()
            return not (a.startswith(("•","ℹ","LÉGENDE","LEGENDE")) or a.upper().startswith("TOTAL"))
        for i, r in enumerate([list(r) for r in wb[name].iter_rows(min_row=4, values_only=True) if _keep(r)], 1):
            loc = f"{name} L{i+3}"
            g = lambda j: r[j] if j < len(r) else None
            fn = str(g(0)).strip() if nonempty(g(0)) else ""
            if not fn: E(loc, "nom de fonds vide.")
            elif funds and fn not in funds: E(loc, f"fonds '{fn}' absent de l'onglet '{ncname}' (jointure impossible).")
            if not _pd(g(1)): E(loc, f"date invalide : {g(1)!r} (DD/MM/YYYY).")
            ft = str(g(2)).strip() if nonempty(g(2)) else ""
            if ft not in FLUXT: E(loc, f"type de flux inconnu '{ft}'.")
            if not isnum(g(3)) or (isnum(g(3)) and g(3) < 0): E(loc, f"montant invalide : {g(3)!r} (numérique positif).")
    return iss

def report(iss, header=True):
    errs = [x for x in iss if x[0] == "ERREUR"]; warns = [x for x in iss if x[0] == "AVERT."]
    if header:
        print(f"── Lint : {len(errs)} erreur(s), {len(warns)} avertissement(s) ──")
    for lvl, where, msg in iss:
        print(f"  [{lvl}] {where} : {msg}")
    return len(errs)

def main():
    if len(sys.argv) < 2: raise SystemExit("Usage: python3 lint.py source.xlsx")
    n_err = report(lint(sys.argv[1]))
    if n_err == 0: print("  ✓ Aucun blocage.")
    sys.exit(1 if n_err else 0)

if __name__ == "__main__": main()
