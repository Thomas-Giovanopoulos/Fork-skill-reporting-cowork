#!/usr/bin/env python3
"""Pont Excel -> manifeste (entrée de P1).

Lit l'onglet 'Entités' (id, label, type, suffixe d'onglet) + scanne les onglets
catégorie pour déterminer, par entité, quelles catégories sont présentes.

Usage : python3 excel_to_manifest.py source.xlsx manifest.json \
            [--period-long "T2 2026"] [--period-short "T2-26"] \
            [--date 2026-06-30] [--date-display "30 juin 2026"] [--version v1]

Les métadonnées de période/version ne sont pas dans l'Excel : passées en option
(ou valeurs par défaut), conformément au cadrage P0.
"""
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8"); sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass
import sys, argparse, unicodedata, re
import openpyxl

PREFIX_TO_CAT = {"liq":"liquidites","immo":"immobilier","fin cote":"financier_cote",
                 "non cote":"non_cote","dettes":"dettes"}
CAT_ORDER = ["liquidites","immobilier","financier_cote","non_cote","dettes"]

def norm(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii","ignore").decode()
    return re.sub(r"\s+"," ", s.lower()).strip()

def sheet_has_data(ws, header_row=3):
    """Vrai seulement si une ligne de DONNÉES réelle existe (col A remplie, hors
    lignes d'aide LÉGENDE/puces/ℹ/TOTAL) — évite de compter un onglet vide comme rempli."""
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        if not row: continue
        a = str(row[0]).strip() if row[0] is not None else ""
        if not a: continue
        if a.startswith(("•", "\u2022", "ℹ", "LÉGENDE", "LEGENDE")) or a.upper().startswith("TOTAL"):
            continue
        return True
    return False

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx"); ap.add_argument("out")
    ap.add_argument("--client"); ap.add_argument("--period-long", default="T1 2026")
    ap.add_argument("--period-short", default="T1-26"); ap.add_argument("--date", default="2026-03-31")
    ap.add_argument("--date-display", default="31 mars 2026"); ap.add_argument("--version", default="v1")
    ap.add_argument("--quarter", choices=["T1","T2","T3","T4"], help="Dérive automatiquement période et date de clôture (T4 = année complète).")
    ap.add_argument("--profile", choices=["Prudent","Équilibré","Dynamique"], help="Profil de risque du client (sinon lu dans l'onglet Entités, défaut Équilibré).")
    ap.add_argument("--show-benchmark", action="store_true", help="Afficher la courbe benchmark du profil (défaut : masqué dans la version présentée).")
    ap.add_argument("--mode", choices=["presentee","envoyee"], default="presentee", help="Fork de reporting : presentee (dynamique, défaut) | envoyee (WIP, statique/simplifié).")
    ap.add_argument("--show-ps-corrige", action="store_true", help="Afficher la courbe coté revalorisée (produits structurés en valeur proratisée). Défaut masqué ; visible seulement si des PS existent.")
    ap.add_argument("--year", type=int, help="Année du reporting (avec --quarter).")
    a = ap.parse_args()

    # --quarter : dérive period_long/short, date de clôture, affichage (override des options explicites)
    if a.quarter:
        import datetime as _dt
        yr = a.year or _dt.date.today().year
        QEND = {"T1":(3,31,"31 mars"),"T2":(6,30,"30 juin"),"T3":(9,30,"30 septembre"),"T4":(12,31,"31 décembre")}
        m,d,disp = QEND[a.quarter]
        a.period_long  = f"Année {yr}" if a.quarter=="T4" else f"{a.quarter} {yr}"   # T4 = année complète
        a.period_short = f"{a.quarter}-{str(yr)[2:]}"
        a.date         = f"{yr}-{m:02d}-{d:02d}"
        a.date_display = f"{disp} {yr}"

    wb = openpyxl.load_workbook(a.xlsx, data_only=True)
    if "Entités" not in wb.sheetnames:
        raise SystemExit("Onglet 'Entités' manquant : impossible de déduire libellés et types.")
    ews = wb["Entités"]
    ents = []
    for row in ews.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]: continue
        eid, label, etype, suffix = (list(row)+[None]*4)[:4]
        ents.append({"id":str(eid).strip(), "label":str(label).strip(),
                     "type":str(etype).strip(), "_suffix":str(suffix).strip()})
    if not ents: raise SystemExit("Onglet 'Entités' vide.")

    # index des onglets catégorie : (suffixe_norm, cat) -> a des données ?
    catsheets = {}
    for name in wb.sheetnames:
        if " — " not in name: continue
        prefix, suffix = name.split(" — ",1)
        cat = PREFIX_TO_CAT.get(norm(prefix))
        if cat: catsheets[(norm(suffix), cat)] = sheet_has_data(wb[name])

    for e in ents:
        sufn = norm(e["_suffix"])
        cats = [c for c in CAT_ORDER if catsheets.get((sufn, c))]
        e["categories"] = cats
        del e["_suffix"]

    _has_nc = any("non_cote" in e["categories"] for e in ents)
    # profil de risque : option CLI > cellule "Profil" de l'onglet Entités > défaut Équilibré
    prof = a.profile
    if not prof and "Entités" in wb.sheetnames:
        for row in wb["Entités"].iter_rows(values_only=True):
            for i,c in enumerate(row or []):
                if isinstance(c,str) and "profil" in c.lower():
                    if ":" in c: prof=c.split(":",1)[1].strip()
                    elif i+1<len(row) and row[i+1]: prof=str(row[i+1]).strip()
                    break
            if prof: break
    pl=(prof or "").lower()
    prof = "Prudent" if pl.startswith("prud") else "Dynamique" if pl.startswith("dynam") else "Équilibré"
    _mode = getattr(a,"mode","presentee")
    client = a.client or next((e["label"] for e in ents if e["type"]=="pp"), ents[0]["label"])
    manifest = {
        "manifest_version":"1.0",
        "reporting":{"client_label":client,"subtitle":"Family office - Reporting",
            "period_long":a.period_long,"period_short":a.period_short,
            "date_reporting":a.date,"date_display":a.date_display,"version":a.version,"profile":prof,"show_benchmark":bool(getattr(a,"show_benchmark",False)),"mode":getattr(a,"mode","presentee"),"show_ps_corrige":bool(getattr(a,"show_ps_corrige",False))},
        "blocs_enabled":{"hero":True,"contexte":True,"supervision":True,"performance":True,
            "performance_nc":_has_nc,
            "repartition":(_mode!="presentee"),"exhaustif":(_mode!="presentee"),"footer":True},
        "entities":ents,
    }
    import json
    json.dump(manifest, open(a.out,"w"), ensure_ascii=False, indent=2)
    print(f"OK manifeste: {a.out}")
    for e in ents: print(f"   {e['label']} [{e['type']}] -> {e['categories']}")

if __name__=="__main__": main()
