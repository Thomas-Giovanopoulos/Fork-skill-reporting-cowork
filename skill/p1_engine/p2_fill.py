#!/usr/bin/env python3
"""P2 — remplissage : Excel source + manifeste -> reporting HTML rempli.

Réutilise la MÊME banque que P1 (structure identique) ; ne fait qu'injecter les
données dans les data-slots et générer les lignes de détail. La structure reste
donc déterministe (pilotée par le manifeste + le code), seules les valeurs varient.

Rempli en v1 : Hero (KPIs), Bloc Répartition (tableau global + cards entité),
Bloc Exhaustif (lignes + sous-totaux + totaux, avec regroupement multi-poches).
Laissé vide (jugement/externe) : donuts d'allocation (classification 12 classes),
Bloc Contexte (macro + indices) et widgets narratifs du Bloc Performance.

Usage : python3 p2_fill.py source.xlsx manifest.json sortie.html
"""
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8"); sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass
import sys, os, unicodedata, re, math
from collections import OrderedDict
import openpyxl
from jinja2 import Environment, FileSystemLoader, StrictUndefined
import assemble as A   # réutilise BLOCK_ORDER, compute_layout, frdate, client_label_html, CATEGORY_*

HIST_CARDS_MAX = 5  # Rendement annuel : <= N années affichées (historique + YTD) -> cartes ; au-delà -> bar chart

# ---------- formatage ----------
def num(v):
    if v in (None, ""): return None
    try: return float(v)
    except (ValueError, TypeError): return None

def multnum(v):
    """Parse un multiple ('1,25x' / '1.25×' / 1.25) -> 1.25, sinon None."""
    if v in (None, ""): return None
    if isinstance(v,(int,float)): return float(v)
    t=str(v).strip().lower().replace("×","").replace("x","").replace(" ","").replace(",",".")
    try: return float(t)
    except (ValueError, TypeError): return None

def eur(v):
    if v is None: return "&mdash;"
    n = round(v); s = "−" if n < 0 else ""
    return f"{s}{abs(n):,.0f} €".replace(",", " ")

def meur(v):
    if v is None: return "&mdash;"
    n = round(v/1e6, 2)
    if n == 0: return "&mdash;"
    s = "−" if n < 0 else ""
    return f"{s}{abs(n):.2f}".replace(".", ",") + " M€"

def pct2(v):
    if v is None: return "&mdash;"
    return f"{v:+.2f}".replace(".", ",") + "%"

def badge(v):
    if v is None: return '<span class="b-neu">&mdash;</span>'
    cls = "b-pos" if v > 0 else ("b-neg" if v < 0 else "b-neu")
    return f'<span class="{cls}">{pct2(v)}</span>'

def muted(txt): return f'<span class="v-muted">{txt}</span>' if txt not in (None,"") else '<span class="v-muted">—</span>'
def name_block(n, d): return f'<div class="cname">{n}</div>' + (f'<div class="cdet">{d}</div>' if d else "")
NATURE = {"av":"Assurance-vie","capi":"Contrat de capitalisation","cto":"CTO"}

# ---------- donuts (allocation consolidée) ----------
import json as _json
CLASS_COLORS = {"Actions":"#14324F","Produits structurés":"#3E6188","Obligations":"#6E93B5","Fonds euros":"#B8975A","Court terme":"#D8C6A0","Monétaire":"#D8C6A0","Alternatifs":"#7C6E99","Matières premières":"#C0824F","Crypto":"#D9A441","Private Equity":"#4F7D64","Actions non cotées":"#8E6C88","Dette privée":"#A98A63","Immo non coté":"#7FA07E","Infrastructures":"#5E7A88"}
DEPO_COLORS = {"UBS":"#14324F","Wealins Lux":"#B8975A","Indosuez":"#3E6188","Edmond de Rothschild":"#6E93B5","NC":"#94A2AD","Tilvest":"#7FA07E"}
ENV_COLORS = {"Assurance-vie LU":"#14324F","Assurance-vie FR":"#B8975A","Capitalisation LU":"#3E6188","Capitalisation FR":"#6E93B5","Compte Titres":"#7FA07E","PEA":"#4F7D64","Nominatif pur":"#A98A63"}
GEO_COLORS = {"Europe développée":"#14324F","Amérique du Nord":"#B8975A","Asie-Pacifique":"#3E6188","Émergents":"#6E93B5","International / Monde":"#D8C6A0"}
FALLBACK = ["#14324F","#B8975A","#3E6188","#6E93B5","#D8C6A0","#4F7D64","#C0824F","#7C6E99","#A98A63","#7FA07E","#5E7A88","#D9A441"]
NCOTE_CLASS = {"PE":"Private Equity","Private Equity":"Private Equity","Actions non cotées":"Actions non cotées","Actions non cotées en direct":"Actions non cotées","Titres non cotés":"Actions non cotées","Dette privée":"Dette privée","Immo non coté":"Immo non coté","Immobilier non coté":"Immo non coté","Infrastructures":"Infrastructures"}

def envelope(nature, depo):
    nat=(nature or "").lower(); fr = (str(depo).strip().upper()=="NC")
    if nat=="pea": return "PEA"
    if nat in ("nominatif","nominatif pur","np"): return "Nominatif pur"
    if nat=="cto": return "Compte Titres"
    if nat=="capi": return "Capitalisation FR" if fr else "Capitalisation LU"
    return "Assurance-vie FR" if fr else "Assurance-vie LU"

def _norm(x):
    return " ".join(str(x or "").split()).casefold()

def _n_mvt(t):
    """Type de mouvement -> 0 versement / 1 retrait / 2 frais (None = inconnu)."""
    t=_norm(t)
    if not t: return None
    if t.startswith(("versement","apport","souscription")): return 0
    if t.startswith(("retrait","rachat")): return 1
    if t.startswith("frais"): return 2
    return None

def _signed_eur(s):
    """Préfixe + (gain) / − déjà présent (perte) sur un montant € déjà formaté."""
    if not s or s == "&mdash;": return s
    st = s.strip()
    if st[0] in ("−", "-", "+"): return s
    digits = "".join(ch for ch in st if ch.isdigit())
    if not digits or digits.strip("0") == "": return s  # zéro -> pas de signe
    return "+" + s

def eur_compact(v):
    if v is None: return "&mdash;"
    a=abs(v); sg="−" if v<0 else ""
    if a>=1e6: return f"{sg}{a/1e6:.1f}".replace(".",",")+" M€"
    if a>=1e3: return f"{sg}{a/1e3:.0f} K€"
    return f"{sg}{a:.0f} €"

SRI_BY_CLASS = {"Actions":6,"Produits structurés":4,"Obligations":3,"Fonds euros":2,
                "Monétaire":1,"Court terme":1,"Alternatifs":4,"Matières premières":5,"Crypto":7,
                "Private Equity":6,"Actions non cotées":7,"Dette privée":4,"Immo non coté":3,"Infrastructures":3}

def sri_of(raw_val, cls, default=4):
    """SRI (1-7) : valeur explicite de la colonne si valide, sinon repli par classe."""
    v = num(raw_val)
    if v is not None and 1 <= v <= 7:
        return int(round(v))
    return SRI_BY_CLASS.get(cls, default)

def sri_label(sri):
    """Binning acté : 1-2 Défensif, 3-4 Équilibré, 5 Dynamique, 6-7 Offensif."""
    if sri is None: return None
    if sri <= 2: return "Défensif"
    if sri <= 4: return "Équilibré"
    if sri == 5: return "Dynamique"
    return "Offensif"

def make_double_donut(inner_agg, nc_names):
    """Anneau intérieur = classes en € ; anneau extérieur = 2 arcs Coté / Non coté."""
    total = sum(inner_agg.values())
    if total <= 0: return None
    import math as _m
    items = sorted(inner_agg.items(), key=lambda kv: kv[1], reverse=True)
    legend, inner_labels, inner_vals, inner_colors = [], [], [], []
    fb = 0
    nc_amt = 0.0
    for name, amt in items:
        col = CLASS_COLORS.get(name)
        if not col: col = FALLBACK[fb % len(FALLBACK)]; fb += 1
        pct = round(amt/total*100)
        is_nc = name in nc_names
        if is_nc: nc_amt += amt
        legend.append({"name":name,"color":col,"pct":pct,"eur":eur_compact(amt),"nc":is_nc})
        inner_labels.append(name); inner_vals.append(round(amt)); inner_colors.append(col)
    cote_amt = total - nc_amt
    outer_labels = ["Coté","Non coté"]
    outer_vals   = [round(cote_amt), round(nc_amt)]
    outer_colors = ["#B8975A","#34495E"]
    return {"center": f"{total/1e6:.1f}".replace(".",","),
            "inner_labels_js": _json.dumps(inner_labels, ensure_ascii=False),
            "inner_data_js": _json.dumps(inner_vals),
            "inner_colors_js": _json.dumps(inner_colors),
            "outer_labels_js": _json.dumps(outer_labels, ensure_ascii=False),
            "outer_data_js": _json.dumps(outer_vals),
            "outer_colors_js": _json.dumps(outer_colors),
            "legend": legend,
            "cote_pct": round(cote_amt/total*100), "nc_pct": round(nc_amt/total*100)}

def make_donut(agg, colmap, maxn=7):
    total = sum(agg.values())
    if total <= 0: return None
    items = sorted(agg.items(), key=lambda kv: kv[1], reverse=True)
    if len(items) > maxn:                       # regroupe la traîne dans "Autres" pour éviter des légendes trop hautes
        head = items[:maxn-1]; autres = sum(a for _, a in items[maxn-1:])
        items = head + [("Autres", autres)]
    legend, labels, data_pct, colors = [], [], [], []
    fb = 0
    for name, amt in items:
        col = colmap.get(name)
        if not col: col = FALLBACK[fb % len(FALLBACK)]; fb += 1
        pct = round(amt/total*100)
        legend.append({"name":name,"color":col,"pct":pct,"eur":eur_compact(amt)})
        labels.append(name); data_pct.append(pct); colors.append(col)
    return {"center": f"{total/1e6:.1f}".replace(".",","),
            "legend": legend,
            "labels_js": _json.dumps(labels, ensure_ascii=False),
            "data_js": _json.dumps(data_pct),
            "colors_js": _json.dumps(colors)}


def g(row, i): return row[i] if i < len(row) else None
def s(v): return "" if v in (None,"") else str(v)

# ---------- non coté : dates, multiples, TRI ----------
_MOIS_FR={"janv":1,"jan":1,"févr":2,"fev":2,"fevr":2,"mars":3,"avr":4,"avril":4,"mai":5,
          "juin":6,"juil":7,"juill":7,"août":8,"aout":8,"sept":9,"septembre":9,"oct":10,"octobre":10,
          "nov":11,"novembre":11,"déc":12,"dec":12,"decembre":12,"décembre":12,"janvier":1,"février":2,
          "fevrier":2,"avril":4,"juillet":7}
def pdate(v):
    """Date Excel (datetime, 'DD/MM/YYYY' ou 'mois AAAA' FR) -> tuple (y,m,d), ou None."""
    if v in (None, ""): return None
    if hasattr(v, "year"): return (v.year, v.month, v.day)
    txt=str(v).strip()
    try:
        d, m, y = (int(x) for x in txt[:10].split("/"))
        return (y, m, d)
    except Exception:
        pass
    try:
        # ISO 'AAAA-MM-JJ[ HH:MM:SS]' : cas d'un datetime Excel déjà coercé en str() en amont (s(...))
        y, m, d = (int(x) for x in txt[:10].split("-"))
        return (y, m, d)
    except Exception:
        pass
    import re as _re
    mo=_re.match(r"([A-Za-zÀ-ÿ]+)\.?\s+(\d{4})", txt)   # ex. 'Févr. 2023', 'Mars 2024'
    if mo:
        k=mo.group(1).lower().rstrip(".")
        if k in _MOIS_FR: return (int(mo.group(2)), _MOIS_FR[k], 1)
    return None

def dmy(t):
    return f"{t[2]:02d}/{t[1]:02d}/{t[0]}" if t else "&mdash;"

def _months_between(d1, d2):
    """Mois écoulés (float, base 30j) entre deux tuples (y,m,d) ; >= 0."""
    if not d1 or not d2: return 0.0
    m=(d2[0]-d1[0])*12 + (d2[1]-d1[1]) + (d2[2]-d1[2])/30.0
    return max(0.0, m)

def ps_metrics(r, rep_ymd, prev_statut, prev_ymd):
    """Proratisation d'un produit structuré. Coupon ANNUEL. Voir CDC_PS."""
    nominal=num(g(r,3)) or 0.0
    coupon=num(g(r,4)) or 0.0            # % annuel (ex: 6 pour 6%)
    debut=pdate(g(r,5)); duree=num(g(r,6))
    seuil=num(g(r,7)); niveau=num(g(r,8)); val_act=num(g(r,9)); rappel=num(g(r,11))
    if niveau is not None and seuil is not None: casse=(niveau < seuil)
    else: casse=(val_act is not None)
    protege=not casse
    rappele=(rappel is not None and niveau is not None and niveau>=rappel)
    statut_code=("rappele" if rappele else ("cassee" if casse else "effective"))
    mois=_months_between(debut, rep_ymd); premiere=(mois<=12)
    vm=(val_act if (casse and val_act is not None) else nominal)          # valeur de marché
    if premiere:
        mois_elig=mois
    else:
        if (prev_statut or "").lower()=="rouge":
            mois_elig=_months_between(debut, prev_ymd) if prev_ymd else mois   # gel au dernier point vert
        else:
            mois_elig=mois
    gain=nominal*(coupon/100.0)*(mois_elig/12.0)
    base=nominal if protege else (val_act if val_act is not None else nominal)
    return {"nom":s(g(r,0)).strip() or "Produit structuré","isin":s(g(r,1)).strip(),
            "env":s(g(r,2)).strip(),"nominal":nominal,"coupon":coupon,"duree":duree,
            "seuil":seuil,"niveau":niveau,"rappel":rappel,"casse":casse,"protege":protege,
            "statut_code":statut_code,"mois":mois,"premiere":premiere,"vm":vm,"gain":gain,
            "val_prorat":base+gain,"cls":s(g(r,10)).strip() or "Produits structurés"}

def _sv_pct(x):
    if x is None: return "\u2014"
    d=float(x)-100.0; return ("+" if d>=0 else "\u2212")+f"{abs(d):g} %"

def ps_pitch(pm):
    cp=pm.get("coupon") or 0; prot=pm.get("seuil"); rap=pm.get("rappel"); niv=pm.get("niveau")
    P=[]
    P.append(f"Produit structuré de type autocall versant un coupon conditionnel de <strong>{cp:g}\u00a0%/an</strong>.")
    if prot is not None:
        P.append(f"Le capital est intégralement remboursé à l'échéance tant que le sous-jacent ne recule pas de plus de <strong>{100-prot:g}\u00a0%</strong> (barrière de protection à {prot:g}\u00a0% du niveau initial).")
    if rap is not None:
        P.append(f"Le produit est rappelé par anticipation dès que le sous-jacent atteint <strong>{rap:g}\u00a0%</strong> de son niveau initial : l'investisseur récupère alors son capital et l'ensemble des coupons cumulés.")
    if pm.get("statut_code")=="rappele":
        P.append("À la dernière constatation, la barrière de rappel est franchie : le produit est en position de <strong>remboursement anticipé</strong>.")
    elif pm.get("casse"):
        P.append(f"À ce jour, le sous-jacent ({_sv_pct(niv)}) est <strong>sous la barrière de protection</strong> : une moins-value latente est retenue dans la valorisation.")
    elif niv is not None:
        P.append(f"À ce jour, le sous-jacent est à {_sv_pct(niv)} de son niveau initial et la <strong>protection reste effective</strong>.")
    return " ".join(P)

def ps_diagram(pm):
    """Petit schéma mécanisme (chronologie + barrières protection/rappel + cours)."""
    W,H=308,144; x0,x1=44,286; yt,yb=12,108
    _vals=[100.0]
    if pm["niveau"] is not None: _vals.append(float(pm["niveau"]))
    if pm["seuil"] is not None: _vals.append(float(pm["seuil"]))
    if pm.get("rappel") is not None: _vals.append(float(pm["rappel"]))
    for _hm,_hn in (pm.get("hist") or []): _vals.append(float(_hn))
    lo=min(_vals); hi=max(_vals); _pad=max(8.0,(hi-lo)*0.12); lo-=_pad; hi+=_pad
    if hi-lo<1: hi=lo+1
    def ymap(L):
        L=max(lo,min(hi,L)); return yb-(L-lo)/(hi-lo)*(yb-yt)
    duree_m=((pm["duree"] or 5)*12.0) or 60.0
    mois=min(pm["mois"], duree_m)
    def xmap(m): return x0+(m/duree_m)*(x1-x0)
    niveau=pm["niveau"] if pm["niveau"] is not None else 100.0
    seuil=pm["seuil"] if pm["seuil"] is not None else 60.0
    rappel=pm.get("rappel")
    y100=ymap(100.0)
    P=[f'<svg viewBox="0 0 {W} {H}" class="ps-diag" xmlns="http://www.w3.org/2000/svg" role="img">']
    import re as _re
    _uid=_re.sub(r"\W","",(pm.get("isin") or pm.get("nom") or "ps"))[:16] or "ps"
    P.append(f'<defs><linearGradient id="psg_{_uid}" x1="0" y1="0" x2="0" y2="1"><stop offset="0" stop-color="#0D1B2A" stop-opacity="0.16"/><stop offset="1" stop-color="#0D1B2A" stop-opacity="0"/></linearGradient></defs>')
    P.append(f'<line x1="{x0}" y1="{y100:.1f}" x2="{x1}" y2="{y100:.1f}" stroke="rgba(13,27,42,.22)" stroke-width="1" stroke-dasharray="2 2"/>')
    yp=ymap(seuil)
    P.append(f'<line x1="{x0}" y1="{yp:.1f}" x2="{x1}" y2="{yp:.1f}" stroke="#B04A3A" stroke-width="1.2" stroke-dasharray="5 3"/>')
    if rappel is not None:
        yr=ymap(rappel)
        P.append(f'<line x1="{x0}" y1="{yr:.1f}" x2="{x1}" y2="{yr:.1f}" stroke="#4F7D64" stroke-width="1.2" stroke-dasharray="5 3"/>')
    ny=int(round(pm["duree"] or 5))
    for k in range(1,ny+1):
        xk=xmap(k*12); P.append(f'<circle cx="{xk:.1f}" cy="{y100:.1f}" r="1.7" fill="#B8975A"/>')
    xn=xmap(mois); yn=ymap(niveau)
    _pts=[(0.0,100.0)]
    for _hm,_hn in (pm.get("hist") or []):
        if 0.0<=_hm<=mois+0.5: _pts.append((_hm,float(_hn)))
    if abs(_pts[-1][0]-mois)>0.5 or abs(_pts[-1][1]-niveau)>1e-6: _pts.append((mois,niveau))
    _xy=[(xmap(_m),ymap(_n)) for (_m,_n) in _pts]
    _area="M"+f"{_xy[0][0]:.1f},{_xy[0][1]:.1f}"+"".join(f" L{px:.1f},{py:.1f}" for px,py in _xy[1:])+f" L{_xy[-1][0]:.1f},{yb:.1f} L{_xy[0][0]:.1f},{yb:.1f} Z"
    P.append(f'<path d="{_area}" fill="url(#psg_{_uid})" stroke="none"/>')
    _dcur="M"+" L".join(f"{px:.1f},{py:.1f}" for px,py in _xy)
    P.append(f'<path d="{_dcur}" fill="none" stroke="#0D1B2A" stroke-width="2"/>')
    for px,py in _xy[1:-1]:
        P.append(f'<circle cx="{px:.1f}" cy="{py:.1f}" r="1.5" fill="#0D1B2A" opacity=".5"/>')
    col=("#4F7D64" if pm.get("statut_code")=="rappele" else ("#B8975A" if pm["protege"] else "#B04A3A"))
    P.append(f'<circle cx="{xn:.1f}" cy="{yn:.1f}" r="3.6" fill="{col}"/>')
    P.append(f'<line x1="{x0}" y1="{yb+4:.1f}" x2="{x1}" y2="{yb+4:.1f}" stroke="rgba(13,27,42,.35)" stroke-width="1"/>')
    for _k in range(0, ny+1):
        _xk=xmap(_k*12)
        P.append(f'<line x1="{_xk:.1f}" y1="{yb+4:.1f}" x2="{_xk:.1f}" y2="{yb+7:.1f}" stroke="rgba(13,27,42,.35)" stroke-width="1"/>')
        if _k==0: _al,_aa,_ax="Émission","start",x0
        elif _k==ny: _al,_aa,_ax="Échéance","middle",x1
        elif ny<=7 or _k%2==0: _al,_aa,_ax=(f"{_k} an"+("s" if _k>1 else "")),"middle",_xk
        else: _al=None
        if _al: P.append(f'<text x="{_ax:.1f}" y="{yb+16:.1f}" class="ps-diag-lbl" text-anchor="{_aa}">{_al}</text>')
    P.append('</svg>')
    return "".join(P)


def mult(v):
    if v is None: return "&mdash;"
    return f"{v:.2f}".replace(".", ",") + "\u00d7"

def tri_fmt(v):
    if v is None: return "&mdash;"
    return f"{v:.1f}".replace(".", ",") + "%"

FLUX_TYPES = {"appel": "Appel", "distribution": "Distribution", "valorisation": "Valorisation",
              "appel prevu": "Appel pr\u00e9vu", "distribution prevue": "Distribution pr\u00e9vue"}
def flux_type(v):
    n = unicodedata.normalize("NFKD", s(v)).encode("ascii", "ignore").decode().lower().strip()
    return FLUX_TYPES.get(n)

# ---------- lecture d'un onglet ----------
def _is_data_row(r):
    """Ligne de données réelle : intitulé (col A) présent et non ligne d'aide du template
    (LÉGENDE, puces, note ℹ, TOTAL). Écarte aussi les lignes vides/placeholder."""
    if not r: return False
    a = str(r[0]).strip() if r[0] is not None else ""
    if not a: return False
    if a.startswith(("•", "\u2022", "ℹ", "LÉGENDE", "LEGENDE")) or a.upper().startswith("TOTAL"):
        return False
    return True

import colmap as _cm

def read_sheet(wb, prefix, suffix):
    name = f"{prefix} — {suffix}"
    if name not in wb.sheetnames: return []
    ws = wb[name]
    mapping = None
    if prefix in _cm.MAPS:
        hdr = [c.value for c in ws[3]]
        try: mapping = _cm.build_mapping(hdr, prefix)
        except Exception: mapping = None
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not _is_data_row(r): continue
        rows.append(_cm.project(list(r), mapping, prefix) if mapping else list(r))
    return rows

# ---------- construction des lignes par catégorie ----------
def rows_liquidites(raw):
    out, tot = [], 0
    for r in raw:
        solde = num(g(r,2)) or 0; tot += solde
        out.append({"tr_class":"rc","td0_style":"","cells":[name_block(s(g(r,0)), s(g(r,1))), eur(num(g(r,2)))]})
    return out, tot

def rows_immobilier(raw):
    out, tot = [], 0
    for r in raw:
        val = num(g(r,7)) or 0; tot += val
        out.append({"tr_class":"rc","td0_style":"","cells":[
            name_block(s(g(r,0)), s(g(r,1))),
            f'<div class="cdet">{s(g(r,2))}</div>', muted(s(g(r,3))), muted(s(g(r,4))),
            muted(s(g(r,5))), muted(eur(num(g(r,6))) ), eur(num(g(r,7)))]})
    return out, tot

def rows_non_cote(raw):
    out, tot = [], 0
    for r in raw:
        eng,app = num(g(r,4)), num(g(r,5)); vl = num(g(r,9)); vl = vl if vl is not None else app
        nar = (eng-app) if (eng is not None and app is not None) else None
        tot += (vl or 0)
        out.append({"tr_class":"rc","td0_style":"","cells":[
            name_block(s(g(r,0)), s(g(r,1))), eur(eng), eur(app), eur(nar),
            f'<span class="v-muted">{eur(num(g(r,7)))}</span>', s(g(r,8)) or "&mdash;", eur(vl)]})
    return out, tot

def rows_dettes(raw):
    out, tot = [], 0
    for r in raw:
        cap = num(g(r,9)); cap = cap if cap is not None else num(g(r,4)); cap = cap or 0; tot += cap
        out.append({"tr_class":"rc","td0_style":"","cells":[
            name_block(s(g(r,0)), s(g(r,1))), muted(s(g(r,3))), eur(num(g(r,4))), s(g(r,5)) or "&mdash;",
            muted(s(g(r,6))), muted(s(g(r,7))), muted(s(g(r,8))),
            f'<span class="v-neg">{eur(cap)}</span>']})
    return out, tot

def fin_cells(cname, cdet, dateinv, nant, nominal, gain, gainpct, gainytd, val, bold=False, indent=False):
    nb = '<span class="tag tag-gold">Oui</span>' if str(nant).lower()=="oui" else '<span class="v-muted">—</span>'
    gain_html = (f'<span class="v-pos">{eur(gain)}</span>' if (gain or 0)>0 else (f'<span class="v-neg">{eur(gain)}</span>' if (gain or 0)<0 else eur(gain)))
    valhtml = f'<span style="font-weight:600">{eur(val)}</span>' if bold else eur(val)
    return {"tr_class":"rc","td0_style":("padding-left:22px;" if indent else ""),"cells":[
        name_block((f'<span style="font-weight:600">{cname}</span>' if bold else cname), cdet),
        muted(dateinv), nb, eur(nominal), gain_html, badge(gainpct), (badge(gainytd) if gainytd is not None else '<span class="v-muted">—</span>'), valhtml]}

def rows_financier_cote(raw):
    # regroupement par (Nature, Assureur, Intermédiaire)
    groups = OrderedDict()
    for r in raw:
        key = (s(g(r,0)), s(g(r,1)), s(g(r,2)))
        groups.setdefault(key, []).append(r)
    out, tot = [], 0
    for (nature, assureur, interm), grp in groups.items():
        nat_full = NATURE.get(nature.lower(), nature)
        if len(grp) == 1:
            r = grp[0]
            nominal, v01, val = num(g(r,10)), num(g(r,11)), num(g(r,12))
            gain = (val-nominal) if (val is not None and nominal is not None) else None
            gpct = (gain/nominal*100) if (gain is not None and nominal) else None
            gytd = ((val-v01)/v01*100) if (val is not None and v01) else None
            tot += (val or 0)
            cdet = " · ".join([x for x in [interm, s(g(r,6)), s(g(r,7))] if x])
            out.append(fin_cells(f"{nat_full} — {assureur}", cdet, s(g(r,8)), s(g(r,9)), nominal, gain, gpct, gytd, val, bold=True))
        else:
            snom=sv01=sval=0; has01=False; nant_any=False
            for r in grp:
                snom += num(g(r,10)) or 0
                if num(g(r,11)) is not None: sv01 += num(g(r,11)); has01=True
                sval += num(g(r,12)) or 0
                if str(g(r,9)).lower()=="oui": nant_any=True
            gain=sval-snom; gpct=(gain/snom*100) if snom else None
            gytd=((sval-sv01)/sv01*100) if (has01 and sv01) else None
            tot += sval
            out.append(fin_cells(f"{nat_full} — {assureur}", f"{interm} · {len(grp)} poches",
                                  "", "Oui" if nant_any else "—", snom, gain, gpct, gytd, sval, bold=True))
            for r in grp:
                nominal,v01,val = num(g(r,10)),num(g(r,11)),num(g(r,12))
                gain=(val-nominal) if (val is not None and nominal is not None) else None
                gpct=(gain/nominal*100) if (gain is not None and nominal) else None
                gytd=((val-v01)/v01*100) if (val is not None and v01) else None
                poche = " · ".join([x for x in [s(g(r,3)), s(g(r,4))] if x])
                cdet = " · ".join([x for x in [s(g(r,5)), s(g(r,6)), s(g(r,7))] if x])
                out.append(fin_cells(f"↳ {poche}", cdet, s(g(r,8)), s(g(r,9)), nominal, gain, gpct, gytd, val, indent=True))
    return out, tot

BUILDERS = {"liquidites":rows_liquidites,"immobilier":rows_immobilier,
            "financier_cote":rows_financier_cote,"non_cote":rows_non_cote,"dettes":rows_dettes}

def main():
    if len(sys.argv) < 4: raise SystemExit("Usage: python3 p2_fill.py source.xlsx manifest.json sortie.html")
    xlsx, manp, outp = sys.argv[1], sys.argv[2], sys.argv[3]
    ctx_path = sys.argv[4] if len(sys.argv) > 4 else None
    manifest = A.load_json(manp)
    A.validate(manifest, A.load_json(A.ROOT/"manifest.schema.json"))
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    try:
        import lint as _lint
        _iss=_lint.lint(xlsx); _ne=sum(1 for x in _iss if x[0]=="ERREUR")
        if _iss:
            print(f"── Lint source : {_ne} erreur(s), {len(_iss)-_ne} avert. ──")
            for lvl,w,m in _iss[:25]: print(f"  [{lvl}] {w} : {m}")
            if _ne: print("  ⚠ Erreurs de saisie : le reporting peut être faux.")
    except Exception:
        pass
    # suffixe d'onglet par entité (depuis l'onglet Entités)
    suf = {}
    ews = wb["Entités"]
    for r in ews.iter_rows(min_row=4, values_only=True):
        if r and r[0]: suf[str(r[0]).strip()] = str(r[3]).strip()

    # --- Produits structurés : statut N-1 (auto via snapshots) + accumulateurs ---
    import os as _os0, glob as _glob0
    _ppid0 = next((e["id"] for e in manifest["entities"] if e["type"]=="pp"), manifest["entities"][0]["id"])
    _curdate0 = manifest["reporting"]["date_reporting"]
    _prev_ps={}; _prev_ps_date=None
    for _f0 in _glob0.glob(_os0.path.join("snapshots", f"{_ppid0}_*.json")):
        try: _sj0=_json.load(open(_f0,encoding="utf-8"))
        except Exception: continue
        if _sj0.get("date_reporting","")<_curdate0 and (_prev_ps_date is None or _sj0["date_reporting"]>_prev_ps_date):
            _prev_ps=_sj0.get("ps_status",{}) or {}; _prev_ps_date=_sj0.get("date_reporting")
    _prev_ymd=tuple(int(x) for x in _prev_ps_date.split("-")) if _prev_ps_date else None
    _rep_ymd=tuple(int(x) for x in _curdate0.split("-"))
    ps_list=[]; ps_vm_total=0.0; ps_prorat_total=0.0; ps_status_cur={}

    data = {"rows":{}, "subtotals":{}, "entity_totals":{}, "synth":{}, "hero":{}, "cards":{}, "donuts":{}, "supervision":{}, "perf":{}, "contexte":{}, "compare_n1":{}}
    cat_brut_global = {c:0 for c in A.CATEGORY_ORDER if c!="dettes"}
    debt_by_cat_global = {}
    tot_assets = tot_debt = 0
    g_v01 = g_val = 0  # base perf
    g_flow = 0          # versements - retraits des contrats comptés en base v01 (perf YTD à périmètre constant)
    g_dietz_w = 0       # Σ flux·poids temporel (Modified Dietz) — ajuste le dénominateur du KPI YTD
    agg_cls={}; agg_geo={}; agg_dep={}; agg_env={}; agg_part={}
    sri_num=0.0; sri_den=0.0; sri_num_cote=0.0; sri_den_cote=0.0
    agg_nc_cls={}  # non coté par classe (PE / Dette privée / Immo non coté / Infra) -> benchmark de pairs
    g_nom=0.0; g_valnow=0.0; nc_total=0.0
    pos_raw=[]    # positions financier coté (une par poche/ligne) : (eid, row)
    nc_funds=[]   # positions non coté (une par fonds) : {name, gest, eng, app, vl, cls, tri, eid}
    nc_flux={}    # nom de fonds -> [{d:(y,m,d), t:type canonique, m:montant}]
    cards_raw={}

    # ---- schéma source v2 : Lignes porteuses de Classe/Géographie/SRI par actif ----
    # Si un contrat possède des lignes classées, les donuts classes/géo et le SRI pondéré
    # se construisent depuis ses lignes ; sinon repli legacy sur la ligne Fin coté
    # (« Classe dominante »). Détection par contrat -> bi-format sans migration.
    _lcls = {}          # eid -> ck_norm -> [{v, cls, geo, sri}]
    _cls_done = set()   # (eid, ck_norm) déjà agrégés
    _geo_from_lines = False
    for _e0 in manifest["entities"]:
        _m0 = {}
        for _r0 in read_sheet(wb, "Lignes", suf.get(_e0["id"], "")):
            _ck0 = _norm(g(_r0, 0)); _lb0 = s(g(_r0, 1)).strip()
            if not _ck0 or not _lb0: continue
            _c0 = s(g(_r0, 6)).strip()
            if not _c0: continue
            _m0.setdefault(_ck0, []).append({"v": num(g(_r0, 3)) or 0.0, "cls": _c0,
                                             "geo": s(g(_r0, 7)).strip(), "sri": g(_r0, 8)})
        _lcls[_e0["id"]] = _m0

    # ---- Mouvements par entité : Versement / Retrait / Frais par clé contrat ----
    _mvts = {}        # eid -> ck_norm -> [vers, rach, frais]
    _mvts_dated = {}  # eid -> ck_norm -> [(date_tuple, flux signé : versement +, retrait −)] — pour Modified Dietz
    for _e1 in manifest["entities"]:
        _mm = {}; _md = {}
        for _r1 in read_sheet(wb, "Mouvements", suf.get(_e1["id"], "")):
            _ck1=_norm(g(_r1,1)); _ty=_n_mvt(s(g(_r1,2))); _mt=num(g(_r1,3))
            if not _ck1 or _ty is None or _mt is None: continue
            _mm.setdefault(_ck1,[0.0,0.0,0.0])[_ty]+=abs(_mt)
            _d1=pdate(g(_r1,0))
            if _d1 and _ty in (0,1):
                _md.setdefault(_ck1,[]).append((_d1, abs(_mt) if _ty==0 else -abs(_mt)))
        _mvts[_e1["id"]]=_mm; _mvts_dated[_e1["id"]]=_md

    for e in manifest["entities"]:
        eid = e["id"]; sx = suf.get(eid, "")
        data["rows"][eid]={}; data["subtotals"][eid]={}
        assets_by_cat={}; debt_by_cat={}; debt_total=0
        for cat in e["categories"]:
            prefix = {"liquidites":"Liq","immobilier":"Immo","financier_cote":"Fin coté","non_cote":"Non coté","dettes":"Dettes"}[cat]
            raw = read_sheet(wb, prefix, sx)
            rows, total = BUILDERS[cat](raw)
            data["rows"][eid][cat]=rows
            data["subtotals"][eid][cat]=eur(-total if cat=="dettes" else total) if total else eur(0)
            if cat=="dettes":
                for r in raw:
                    amt = num(g(r,9)); amt = amt if amt is not None else num(g(r,4)); amt = amt or 0
                    a = s(g(r,10)).strip()
                    if a not in A.CATEGORY_ORDER or a=="dettes": a="_non_affectee"
                    debt_by_cat[a]=debt_by_cat.get(a,0)+amt; debt_total += amt
            else:
                assets_by_cat[cat]=total; cat_brut_global[cat]+=total
            if cat=="financier_cote":
                for r in raw:
                    pos_raw.append((eid, r))
                    v01=num(g(r,11)); val=num(g(r,12)) or 0
                    if v01 is not None: g_v01+=v01; g_val += (val if val else v01)
                    else:
                        # KPI sans valeur 01/01 : contrat ouvert dans l'année -> base = capital investi net
                        _di=pdate(g(r,8))
                        _repy=int(manifest["reporting"]["date_reporting"][:4])
                        if _di and _di[0]==_repy:
                            _cap=(num(g(r,10)) or 0)+(num(g(r,17)) or 0)-(num(g(r,18)) or 0)-(num(g(r,19)) or 0)
                            if _cap>0: g_v01+=_cap; g_val+=val
                    nm=num(g(r,10));
                    if nm is not None: g_nom+=nm; g_valnow+=val
                    cls=s(g(r,13)) or "Actions"; geo=s(g(r,14)); depo=s(g(r,5)) or "NC"
                    # clé contrat (mêmes variantes que le Widget 4) -> lignes classées ?
                    _cknat=NATURE.get(s(g(r,0)).lower(), s(g(r,0)))
                    # Mouvements -> vers/rach/frais si les colonnes ne sont pas déjà saisies
                    for _ckm in (_norm(f"{_cknat} — {s(g(r,1))}"), _norm(f"{_cknat} \u00b7 {s(g(r,1))}")):
                        _mv=_mvts.get(eid,{}).pop(_ckm,None)
                        if _mv and not any(num(g(r,k)) for k in (17,18,19)):
                            while len(r)<20: r.append(None)
                            r[17],r[18],r[19]=_mv[0] or None,_mv[1] or None,_mv[2] or None
                    if v01 is not None:
                        # flux intra-période : perf YTD à périmètre constant (Modified Dietz).
                        # SEULS les flux de l'ANNÉE DU REPORTING comptent ici — les Mouvements peuvent
                        # porter des flux d'années antérieures (servant au Dietz depuis origine).
                        import datetime as _dtz
                        _repy2=int(manifest["reporting"]["date_reporting"][:4])
                        _y0=_dtz.date(_repy2,1,1)
                        _dr=_dtz.date(*(int(x) for x in manifest["reporting"]["date_reporting"].split("-")))
                        _T=(_dr-_y0).days or 1
                        _dfl2=[f for _ckm2 in (_norm(f"{_cknat} — {s(g(r,1))}"), _norm(f"{_cknat} \u00b7 {s(g(r,1))}"))
                               for f in _mvts_dated.get(eid,{}).get(_ckm2,[])]
                        if _dfl2:
                            for (_fd,_fa) in _dfl2:
                                if _fd[0]==_repy2:
                                    _t=max(0,(_dtz.date(*_fd)-_y0).days)
                                    g_flow += _fa; g_dietz_w += _fa*max(0.0,(_T-_t)/_T)
                        else:
                            # pas de flux datés : les colonnes vers/rach saisies directement sont réputées de l'année
                            g_flow += (num(g(r,17)) or 0) - (num(g(r,18)) or 0)
                    _ckv={_norm(f"{_cknat} — {s(g(r,1))}"), _norm(f"{_cknat} \u00b7 {s(g(r,1))}")}
                    _hit=[k for k in _ckv if k in _lcls.get(eid,{})]
                    if _hit:
                        for _ckn in _hit:
                            if (eid,_ckn) in _cls_done: continue
                            _cls_done.add((eid,_ckn)); _geo_from_lines=True
                            for _L in _lcls[eid][_ckn]:
                                _lv=_L["v"]; _lc=_L["cls"]
                                agg_cls[_lc]=agg_cls.get(_lc,0)+_lv
                                _ls=sri_of(_L["sri"], _lc)
                                if _lv:
                                    sri_num+=_ls*_lv; sri_den+=_lv
                                    sri_num_cote+=_ls*_lv; sri_den_cote+=_lv
                                if _L["geo"]: agg_geo[_L["geo"]]=agg_geo.get(_L["geo"],0)+_lv
                    else:
                        agg_cls[cls]=agg_cls.get(cls,0)+val
                        _sri=sri_of(g(r,15), cls)
                        if val:
                            sri_num+=_sri*val; sri_den+=val
                            sri_num_cote+=_sri*val; sri_den_cote+=val
                        if cls=="Actions" and geo: agg_geo[geo]=agg_geo.get(geo,0)+val
                    agg_dep[depo]=agg_dep.get(depo,0)+val
                    _pc=s(g(r,1)).strip() or "NC"; agg_part[_pc]=agg_part.get(_pc,0)+val
                    env=envelope(g(r,0), depo); agg_env[env]=agg_env.get(env,0)+val
            if cat=="non_cote":
                for r in raw:
                    vl=num(g(r,9)); vl=vl if vl is not None else num(g(r,5)); vl=vl or 0
                    ncls=NCOTE_CLASS.get(s(g(r,10)), s(g(r,10)) or "Private Equity")
                    agg_cls[ncls]=agg_cls.get(ncls,0)+vl; nc_total+=vl
                    agg_nc_cls[ncls]=agg_nc_cls.get(ncls,0)+vl
                    _part=s(g(r,1)).strip() or "NC"; agg_part[_part]=agg_part.get(_part,0)+vl
                    _sri_nc=sri_of(None, ncls, default=5)
                    if vl: sri_num+=_sri_nc*vl; sri_den+=vl
                    nc_funds.append({"name":s(g(r,0)).strip(),"gest":s(g(r,1)).strip(),
                                     "strat":s(g(r,2)).strip(),"cible":s(g(r,3)).strip(),
                                     "eng":num(g(r,4)),"app":num(g(r,5)),
                                     "na_reel":num(g(r,6)),"na_est":num(g(r,7)),
                                     "vl":num(g(r,9)),
                                     "cls":ncls,"tri":num(g(r,11)),
                                     "moic_x":multnum(g(r,8)),"cible_x":multnum(g(r,3)),
                                     "segment":s(g(r,12)).strip(),"duree":s(g(r,13)).strip(),
                                     "date_inv":s(g(r,14)).strip(),"typ":s(g(r,15)).strip().lower(),"eid":eid})
                for fr in read_sheet(wb, "NC Flux", sx):
                    fn=s(g(fr,0)).strip(); fd=pdate(g(fr,1)); ft=flux_type(g(fr,2)); fm=num(g(fr,3))
                    if not fn or fd is None or ft is None or fm is None: continue
                    nc_flux.setdefault(fn, []).append({"d":fd,"t":ft,"m":abs(fm)})
        # --- Produits structurés (onglet dédié) réintégrés au financier coté ---
        _coursps={}
        for _cr in read_sheet(wb, "Cours PS", sx):
            _ci=s(g(_cr,1)).strip(); _cd=pdate(g(_cr,0)); _cn=num(g(_cr,2))
            if not _ci or _cd is None or _cn is None: continue
            _coursps.setdefault(_ci, []).append((_cd,_cn))
        for pr in read_sheet(wb, "Produits structurés", sx):
            if num(g(pr,3)) is None: continue
            _isin=s(g(pr,1)).strip()
            _pstat=(_prev_ps.get(_isin) or {}).get("statut")
            pm=ps_metrics(pr, _rep_ymd, _pstat, _prev_ymd); pm["eid"]=eid
            _deb=pdate(g(pr,5))
            pm["hist"]=sorted(((_months_between(_deb,_d),float(_n)) for (_d,_n) in _coursps.get(_isin,[])), key=lambda z:z[0])
            _vm=pm["vm"]
            assets_by_cat["financier_cote"]=assets_by_cat.get("financier_cote",0)+_vm
            cat_brut_global["financier_cote"]+=_vm
            agg_cls["Produits structurés"]=agg_cls.get("Produits structurés",0)+_vm
            _pd=pm["env"] or "NC"
            agg_dep[_pd]=agg_dep.get(_pd,0)+_vm
            agg_part[_pd]=agg_part.get(_pd,0)+_vm
            agg_env[_pd]=agg_env.get(_pd,0)+_vm
            g_v01+=pm["nominal"]; g_val+=_vm; g_nom+=pm["nominal"]; g_valnow+=_vm
            _srips=sri_of(None,"Produits structurés")
            if _vm: sri_num+=_srips*_vm; sri_den+=_vm; sri_num_cote+=_srips*_vm; sri_den_cote+=_vm
            ps_list.append(pm); ps_vm_total+=_vm; ps_prorat_total+=pm["val_prorat"]
            ps_status_cur[_isin or pm["nom"]]={"statut":("rouge" if pm["casse"] else "vert")}
        for c,amt in debt_by_cat.items():
            debt_by_cat_global[c]=debt_by_cat_global.get(c,0)+amt
        # cards : valeurs BRUTES (formatées après choix de l'unité adaptative)
        craw={}
        for cat in A.CATEGORY_ORDER:
            if cat=="dettes":
                craw[cat]={"brut":None,"dettes":(-debt_total if debt_total else None),"net":(-debt_total if debt_total else None)}
            else:
                b=assets_by_cat.get(cat); d=debt_by_cat.get(cat,0)
                if b is None and not d:
                    craw[cat]={"brut":None,"dettes":None,"net":None}
                else:
                    bb=b or 0
                    craw[cat]={"brut":(bb if b is not None else None),"dettes":(-d if d else None),"net":(bb-d)}
        assets=sum(assets_by_cat.values()); net=assets-debt_total
        craw["total"]={"brut":assets,"dettes":(-debt_total if debt_total else None),"net":net}
        cards_raw[eid]=craw
        data["entity_totals"][eid]=eur(net)
        tot_assets += assets; tot_debt += debt_total

    # unité adaptative (cards + courbe) selon l'ordre de grandeur du patrimoine
    U_SCALE,U_SUFFIX,U_DEC = (1e6,"M€",2) if tot_assets>=1e6 else (1e3,"K€",0)
    def funit(v):
        if v is None: return "&mdash;"
        n=round(v/U_SCALE,U_DEC)
        if n==0: return "&mdash;"
        sgn="−" if n<0 else ""
        return f"{sgn}{abs(n):.{U_DEC}f}".replace(".",",")+f" {U_SUFFIX}"
    # échelle ADAPTATIVE PAR ENTITÉ (chaque entité dans son unité selon son propre total brut),
    # au lieu de l'unité globale du patrimoine -> une holding à 30 k€ s'affiche "30 K€" et non "0,03 M€".
    def funit_e(v, scale, suf, dec):
        if v is None: return "&mdash;"
        n=round(v/scale,dec)
        if n==0: return "&mdash;"
        sgn="−" if n<0 else ""
        return f"{sgn}{abs(n):.{dec}f}".replace(".",",")+f" {suf}"
    for eid,craw in cards_raw.items():
        _etot=abs(craw.get("total",{}).get("brut") or 0)
        _sc,_su,_de = (1e6,"M€",2) if _etot>=1e6 else (1e3,"K€",0)
        data["cards"][eid]={k:{kk:funit_e(vv,_sc,_su,_de) for kk,vv in row.items()} for k,row in craw.items()}

    # tableau global (dettes réparties par catégorie)
    for cat in [c for c in A.CATEGORY_ORDER if c!="dettes"]:
        brut=cat_brut_global[cat]; d=debt_by_cat_global.get(cat,0); net=brut-d
        if not brut and not d:
            data["synth"][cat]={"brut":"&mdash;","dettes":"&mdash;","net":"&mdash;","pct":"&mdash;"}
        else:
            data["synth"][cat]={"brut":(eur(brut) if brut else "&mdash;"),"dettes":(eur(-d) if d else "&mdash;"),
                                "net":eur(net),"pct":(f"{brut/tot_assets*100:.1f}%".replace(".",",") if tot_assets else "&mdash;")}
    data["synth"]["dettes"]=({"brut":"&mdash;","dettes":eur(-tot_debt),"net":eur(-tot_debt),"pct":"&mdash;"}
                             if tot_debt else {"brut":"&mdash;","dettes":"&mdash;","net":"&mdash;","pct":"&mdash;"})
    data["synth"]["total"]={"brut":eur(tot_assets),"dettes":(eur(-tot_debt) if tot_debt else "&mdash;"),
                            "net":eur(tot_assets-tot_debt),"pct":""}

    # hero
    gain = (g_val - g_v01 - g_flow) if g_v01 else None   # corrigé des versements/retraits (périmètre constant)
    data["hero"]={"actif_brut":eur(tot_assets),"dettes":eur(-tot_debt),"actif_net":eur(tot_assets-tot_debt),
                  "gain_ytd": (f'{eur(gain)} ({pct2(gain/g_v01*100)})' if g_v01 else "&mdash;"),
                  "gain_ytd_base": (f"base : {eur(g_v01)}" if g_v01 else ""),
                  "gain_cls": ("c-neg" if (gain is not None and gain<0) else "c-pos")}

    # donuts consolidés
    data["geo_title"]="Exposition géographique" if _geo_from_lines else "Géographie — Actions"
    data["geo_empty"]="aucune exposition renseignée" if _geo_from_lines else "aucune action en direct"
    for cid, agg, cm in [("donut_classes",agg_cls,CLASS_COLORS),("donut_geo",agg_geo,GEO_COLORS),
                          ("donut_depositaires",agg_dep,DEPO_COLORS),("donut_enveloppes",agg_env,ENV_COLORS)]:
        d = make_donut(agg, cm)
        if d: data["donuts"][cid]=d

    # ---- Bloc Performance ----
    perf = {}
    pf_value = sum(v for v in agg_cls.values())  # coté + non coté (= total classes)
    ytd_pct = (g_val-g_v01-g_flow)/(g_v01+g_dietz_w)*100 if (g_v01+g_dietz_w) else None   # Modified Dietz
    incep_pct = (g_valnow-g_nom)/g_nom*100 if g_nom else None
    _cote_tot = cat_brut_global.get("financier_cote", 0) or 0
    ytd_label = "Performance YTD"
    _ytd_eur_v = (g_val - g_v01 - g_flow) if g_v01 else None
    ytd_label_cote = "Performance YTD — coté"
    if not g_v01 and g_nom:
        # aucun contrat avec base YTD -> perf depuis origine (base = capitaux investis nets)
        ytd_pct = incep_pct; _ytd_eur_v = g_valnow - g_nom
        g_v01_disp = g_nom; ytd_label = "Perf. depuis origine"
        ytd_label_cote = "Perf. depuis origine — coté"
    else:
        g_v01_disp = g_v01
        if g_v01 and _cote_tot and g_val < 0.995*_cote_tot:
            _covp = f"{g_val/_cote_tot*100:.0f}"
            ytd_label = f"Performance YTD ({_covp} % du coté)"
            ytd_label_cote = f"Perf. YTD — coté ({_covp} % couvert)"
    ytd_cls = "c-neg" if (ytd_pct is not None and ytd_pct<0) else "c-pos"
    ytd_card = "card-neg" if (ytd_pct is not None and ytd_pct<0) else "card-pos"
    incep_cls = "c-neg" if (incep_pct is not None and incep_pct<0) else "c-pos"
    start_value = (g_v01 + nc_total) if (g_v01 or nc_total) else None
    _sri_cote = None
    if sri_den_cote > 0:
        _sri_cote = max(1, min(7, int(math.ceil(sri_num_cote/sri_den_cote))))
    perf["kpis"] = {
        "start_value": eur(start_value) if start_value else "&mdash;",
        "ytd_pct": pct2(ytd_pct),
        "ytd_label": ytd_label,
        "ytd_label_cote": ytd_label_cote,
        "ytd_eur": eur(_ytd_eur_v) if _ytd_eur_v is not None else "&mdash;",
        "ytd_eur_signed": _signed_eur(eur(_ytd_eur_v)) if _ytd_eur_v is not None else "&mdash;",
        "portfolio_value": eur(pf_value),
        "cote_start": eur(g_v01_disp) if g_v01_disp else "&mdash;",
        "cote_end": eur(g_val) if g_val else "&mdash;",
        "profil_cote": sri_label(_sri_cote) if _sri_cote else "&mdash;",
        "sri_cote": _sri_cote,
        "ytd_cls": ytd_cls, "ytd_card": ytd_card,
        "flow_note": (f"après retraits nets de {eur(-g_flow)}" if g_flow < 0
                      else (f"après versements nets de {eur(g_flow)}" if g_flow > 0 else None)),
    }
    # --- Tableau "Historique du patrimoine" (Bloc 01) : 1 ligne par arrêté ---
    _hp_rows=[]
    if "Valorisations" in wb.sheetnames:
        import datetime as _dth
        _ryh=int(manifest["reporting"]["date_reporting"][:4])
        _pts={}; _ncp={}
        for row in wb["Valorisations"].iter_rows(min_row=4, values_only=True):
            if not _is_data_row(row): continue
            try: _dd,_mm,_yy=(int(x) for x in str(row[0]).strip()[:10].split("/"))
            except Exception: continue
            _pts[(_yy,_mm,_dd)]=num(row[1]) or 0
            if num(row[2]) is not None: _ncp[(_yy,_mm,_dd)]=num(row[2])
        def _ncat(k):
            # non coté à la date k : dernier point connu <= k (forward-fill, convention de la courbe)
            _c=[v for kk,v in sorted(_ncp.items()) if kk<=k]
            return _c[-1] if _c else None
        def _allocs(cote,nc):
            if nc is None or not (cote or nc): return "&mdash;","&mdash;","&mdash;"
            pat=cote+nc
            a=f"{round(cote/pat*100)}/{round(nc/pat*100)}" if pat else "&mdash;"
            return eur(nc), eur(pat), a
        _base=_pts.get((_ryh-1,12,31))
        _rept=tuple(int(x) for x in manifest["reporting"]["date_reporting"].split("-"))
        _qs=[k for k in sorted(_pts) if k[0]==_ryh and ((k[1],k[2]) in ((3,31),(6,30),(9,30),(12,31)) or k==_rept)]
        _fl=[(f, a) for _em in _mvts_dated.values() for _fs in _em.values() for (f,a) in _fs]
        def _dz(v0,v1,d0,d1):
            _T=(_dth.date(*d1)-_dth.date(*d0)).days or 1
            _cf=0.0; _wf=0.0
            for (fd,fa) in _fl:
                if d0 < fd <= d1:
                    _t=(_dth.date(*fd)-_dth.date(*d0)).days
                    _cf+=fa; _wf+=fa*max(0.0,(_T-_t)/_T)
            _den=v0+_wf
            return (((v1-v0-_cf)/_den*100) if _den else None), _cf
        if _base and _qs:
            _prev=(_ryh-1,12,31); _pv=_base
            _QL={3:"T1",6:"T2",9:"T3",12:"T4"}
            _qrows=[]
            for _q in _qs:
                _vq=_pts[_q]
                _pq,_cfq=_dz(_pv,_vq,_prev,_q)
                _pc,_=_dz(_base,_vq,(_ryh-1,12,31),_q)
                _lblq=(f"{_QL[_q[1]]} {_ryh}" if (_q[1],_q[2]) in ((3,31),(6,30),(9,30),(12,31)) else f"Au {_q[2]:02d}/{_q[1]:02d}/{_ryh}")
                _ncv,_patv,_alv=_allocs(_vq,_ncat(_q))
                _ge=_vq-_pv-_cfq
                _gcl="v-pos" if _ge>0 else ("v-neg" if _ge<0 else "")
                _qrows.append({"periode":_lblq,"valeur":eur(_vq),"nc":_ncv,"pat":_patv,"alloc":_alv,
                               "flux":(_signed_eur(eur(_cfq)) if _cfq else "&mdash;"),
                               "perf":badge(_pq),
                               "perf_eur":(f'<span class="{_gcl}">{_signed_eur(eur(_ge))}</span>' if _gcl else eur(_ge)),
                               "cum":badge(_pc)})
                _prev=_q; _pv=_vq
            _hp_rows.extend(reversed(_qrows))   # du plus récent au plus ancien
    if "Historique" in wb.sheetnames:
        _hy=[]
        for row in wb["Historique"].iter_rows(min_row=4, values_only=True):
            if not _is_data_row(row): continue
            try: _yr=int(str(row[0]).strip()[:4])
            except Exception: continue
            _pf=num(row[1]); _cm=s(row[2] if len(row)>2 else "").lower()
            if _pf is not None:
                _lbl=f"Année {_yr}" + (" (périmètre partiel)" if "partiel" in _cm else "")
                _vy=_pts.get((_yr,12,31)) if "Valorisations" in wb.sheetnames else None
                _fy=sum(a for (fd,a) in _fl if fd[0]==_yr) if "Valorisations" in wb.sheetnames else 0
                _ncv,_patv,_alv=_allocs(_vy,_ncat((_yr,12,31))) if _vy else ("&mdash;","&mdash;","&mdash;")
                _vprev=_pts.get((_yr-1,12,31))
                _ge=(_vy-_vprev-_fy) if (_vy and _vprev is not None) else None
                _gcl=("v-pos" if _ge>0 else ("v-neg" if _ge<0 else "")) if _ge is not None else ""
                _hy.append({"periode":_lbl,
                            "valeur":(eur(_vy) if _vy else "&mdash;"),"nc":_ncv,"pat":_patv,"alloc":_alv,
                            "flux":(_signed_eur(eur(_fy)) if _fy else "&mdash;"),
                            "perf":badge(_pf*100 if abs(_pf)<1 else _pf),
                            "perf_eur":((f'<span class="{_gcl}">{_signed_eur(eur(_ge))}</span>' if _gcl else eur(_ge)) if _ge is not None else "&mdash;"),
                            "cum":"&mdash;","_yr":_yr})
        for _h in sorted(_hy, key=lambda x:-x["_yr"]):
            _h.pop("_yr",None); _hp_rows.append(_h)
    data["histo_pat"]=_hp_rows

    # courbe d'évolution du PATRIMOINE GLOBAL : financier(t) + actif non-financier net (quasi-constant)
    const_nonfin = (tot_assets - pf_value) - tot_debt  # liquidités + immobilier − dettes
    if "Valorisations" in wb.sheetnames:
        vs = wb["Valorisations"]; last_nc=0.0; _curve_end=None
        _ry,_rm,_rd = (int(x) for x in manifest["reporting"]["date_reporting"].split("-"))
        from collections import OrderedDict as _OD
        _month = _OD()   # (an, mois) -> (jour, label, cote, curve_end) : forward-fill NC, on garde le dernier point du mois
        for row in vs.iter_rows(min_row=4, values_only=True):
            if not _is_data_row(row): continue
            d=str(row[0]).strip(); cote=num(row[1]) or 0; ncv=num(row[2])
            try:
                _dd,_mm,_yy = (int(x) for x in d[:10].split("/"))
            except Exception:
                continue
            if (_yy,_mm,_dd) > (_ry,_rm,_rd): continue
            if ncv is not None: last_nc=ncv
            _curve_end=cote+last_nc+const_nonfin+ps_vm_total
            _lbl=f"{_mm:02d}/{str(_yy)[-2:]}"
            _month[(_yy,_mm)] = (_dd, _lbl, cote, _curve_end)   # points en ordre -> dernier jour = fin de mois
        labels=[]; series=[]; fin_eur=[]
        for (_yy,_mm),(_dd,_lbl,cote,ce) in _month.items():
            labels.append(_lbl); series.append(round(ce/U_SCALE,U_DEC)); fin_eur.append(cote)
        if labels:
            _curve_end = fin_eur and (series[-1]*U_SCALE) or _curve_end
            perf["curve"]={"labels_js":_json.dumps(labels),"data_js":_json.dumps(series),"unit":U_SUFFIX,"dec":U_DEC}
            perf["curve_end_eur"]=_curve_end
    # comparaison vs indices (onglet Indices)
    BAR_FULL_PCT = 40.0  # échelle absolue : 40% YTD remplit la barre (réserve en marché haussier)
    def bar(p):
        cls="perf-pos" if (p or 0)>0 else ("perf-neg" if (p or 0)<0 else "perf-neu")
        barcls="bench-bar-pos" if (p or 0)>=0 else "bench-bar-neg"
        return cls, barcls, round(min(100,abs(p or 0)/BAR_FULL_PCT*100))

    # contexte de marché : store par période (réutilisable), produit une fois par trimestre
    cpath = ctx_path or os.path.join("contexte", f"{manifest['reporting']['period_short']}.json")
    store = _json.load(open(cpath, encoding="utf-8")) if os.path.exists(cpath) else {}
    sidx = store.get("indices", [])
    idx_ytd = {i["name"]: i.get("ytd") for i in sidx}

    # --- Bloc CONTEXTE : panorama technique (variation en points + % YTD), SANS portefeuille ---
    ctx_idx=[]
    for i in sidx:
        p=i.get("ytd"); cls,barcls,w=bar(p)
        ctx_idx.append({"name":i["name"],"var":i.get("var",""),
                        "perf":(pct2(p) if p is not None else "&mdash;"),
                        "cls":cls,"bar_cls":barcls,"width":w})
    data["contexte"]={"indices":ctx_idx,"macro_text":store.get("macro_text",[]),
                      "faits":store.get("faits_marquants",[])}

    # --- Bloc PERFORMANCE : Portefeuille vs benchmark du PROFIL DE RISQUE du client ---
    BENCH_PROFILES={"Prudent":{"MSCI World":25,"Obligations (Agg €)":65,"Monétaire €":10},
                    "Équilibré":{"MSCI World":50,"Obligations (Agg €)":45,"Monétaire €":5},
                    "Dynamique":{"MSCI World":80,"Obligations (Agg €)":18,"Monétaire €":2}}
    blocks = store.get("benchmark_blocks", {})
    prof = manifest["reporting"].get("profile","Équilibré")
    weights = BENCH_PROFILES.get(prof, BENCH_PROFILES["Équilibré"])
    out=[]
    cls,barcls,w=bar(ytd_pct)
    out.append({"name":"Portefeuille","is_port":True,"perf":(pct2(ytd_pct) if ytd_pct is not None else "&mdash;")+" YTD",
                "cls":cls,"bar_cls":barcls,"width":w,"blocks":[]})
    tot=0.0; ok=True; comp=[]
    for bname,wt in weights.items():
        by=blocks.get(bname); ccls,cbar,cw=bar(by)
        comp.append({"name":bname,"weight":wt,"perf":(pct2(by) if by is not None else "&mdash;"),
                     "cls":ccls,"bar_cls":cbar,"width":cw})
        if by is None: ok=False
        else: tot+=wt/100.0*by
    bperf=tot if ok else None
    cl,bc,ww=bar(bperf)
    out.append({"name":f"Benchmark {prof}","is_port":False,"perf":(pct2(bperf) if bperf is not None else "&mdash;")+" YTD",
                "cls":cl,"bar_cls":bc,"width":ww,"blocks":comp})
    perf["compare"]=out
    # texte explicatif du profil
    _LBL={"MSCI World":"actions internationales","Obligations (Agg €)":"obligations","Monétaire €":"monétaire"}
    wsum=" · ".join(f"{wt}% {_LBL.get(bn,bn)}" for bn,wt in weights.items())
    txt=f"Votre profil de risque est <strong>{prof}</strong> — un portefeuille de référence composé de {wsum}."
    if ytd_pct is not None and bperf is not None:
        d=ytd_pct-bperf; v=("surperforme" if d>0.3 else "sous-performe" if d<-0.3 else "est en ligne avec")
        txt+=f" Sur la période, votre portefeuille (<strong>{pct2(ytd_pct)}</strong>) {v} ce benchmark ({pct2(bperf)})."
    perf["profile_text"]=txt
    # ---- page « Détails » du bloc 01 : tableau des positions du portefeuille coté ----
    if pos_raw:
        _elbl={e["id"]:e["label"] for e in manifest["entities"]}
        _multi=len(manifest["entities"])>1
        _pv=[abs(x) for _,r in pos_raw for x in (num(g(r,10)),num(g(r,11)),num(g(r,12))) if x]
        _psc,_pun,_pdc=(1e6,"M€",2) if (_pv and max(_pv)>=1e6) else (1e3,"K€",0)
        def eur_p(v):
            if v is None: return "&mdash;"
            n=round(v/_psc,_pdc)
            if n==0: return "&mdash;"
            sg="−" if n<0 else ""
            return f"{sg}{abs(n):.{_pdc}f}".replace(".",",")+f" {_pun}"
        def _gain_html(gv):
            if gv is None: return "&mdash;"
            c="v-pos" if gv>0 else ("v-neg" if gv<0 else "")
            return f'<span class="{c}">{eur_p(gv)}</span>' if c else eur_p(gv)
        prows=[]; t_nom=t_v01=t_val=0.0; _has01=False
        for eid,r in pos_raw:
            nature=NATURE.get(s(g(r,0)).lower(), s(g(r,0))); assureur=s(g(r,1))
            poche=" · ".join(x for x in [s(g(r,3)), s(g(r,4))] if x)
            det=" · ".join(x for x in [poche, (_elbl.get(eid,"") if _multi else "")] if x)
            nom,v01,val=num(g(r,10)),num(g(r,11)),num(g(r,12))
            gain=(val-v01) if (val is not None and v01) else None
            ppct=((val-v01)/v01*100) if (val is not None and v01) else None
            ipct=((val-nom)/nom*100) if (val is not None and nom) else None
            t_nom+=nom or 0; t_val+=val or 0
            if v01 is not None: t_v01+=v01; _has01=True
            prows.append({"name":f"{nature} — {assureur}","det":det,"depo":s(g(r,5)) or "&mdash;",
                          "nom":eur_p(nom),"v01":eur_p(v01),"val":eur_p(val),
                          "gain":_gain_html(gain),"ppct":badge(ppct),"ipct":badge(ipct)})
        _tg=(t_val-t_v01) if _has01 else None
        perf["positions"]={"rows":prows,"unit":_pun,
            "total":{"nom":eur_p(t_nom or None),"v01":(eur_p(t_v01) if _has01 else "&mdash;"),
                     "val":eur_p(t_val or None),"gain":_gain_html(_tg),
                     "ppct":badge((_tg/t_v01*100) if (_tg is not None and t_v01) else None),
                     "ipct":badge(((t_val-t_nom)/t_nom*100) if t_nom else None)}}
    def _sv(v):   # variation signée en points (niveau/seuil exprimés en % de l'initial)
        d=v-100.0; sg="+" if d>=0 else "−"; return f"{sg}{abs(d):g}%"
    _STL={"effective":("Effective","c-gold"),"cassee":("Cassée","c-neg"),"rappele":("Rappelé","c-pos")}
    _psd=[]
    for _p in ps_list:
        _cp=_p["coupon"]
        _lbl,_cls=_STL.get(_p["statut_code"],("Effective","c-pos"))
        _vocab=""
        if _p["niveau"] is not None and _p["seuil"] is not None:
            _vocab=f"sous-jacent {_sv(_p['niveau'])} · protection {_sv(_p['seuil'])}"
            if _p.get("rappel") is not None: _vocab+=f" · rappel {_sv(_p['rappel'])}"
        _psd.append({
          "nom":_p["nom"], "isin":_p["isin"] or "&mdash;", "env":_p["env"] or "&mdash;",
          "nominal":eur(_p["nominal"]),
          "coupon":((f"{_cp:g}".replace(".",",")+" %/an") if _cp else "&mdash;"),
          "statut":_lbl, "statut_cls":_cls,
          "seuil_niveau":_vocab,
          "vm":eur(_p["vm"]),
          "gain":(_signed_eur(eur(_p["gain"])) if _p["gain"] else eur(0)),
          "prorat":eur(_p["val_prorat"]),
          "svg":ps_diagram(_p),
          "pitch":ps_pitch(_p),
        })
    perf["ps"]=_psd
    _psgain=ps_prorat_total-ps_vm_total
    perf["ps_totals"]={"vm":eur(ps_vm_total),"prorat":eur(ps_prorat_total),
                       "gain":(_signed_eur(eur(_psgain)) if _psgain else eur(0)),"n":len(ps_list)}
    perf["has_ps"]=bool(ps_list)
    data["perf"]=perf
    # Comparaison sur la POCHE FINANCIÈRE : trajectoire réelle + équivalent benchmark (€), même point de départ.
    # Le benchmark de profil ne concerne que le financier coté -> on le rebase sur cette poche (pas le patrimoine global).
    if perf.get("curve") and fin_eur:
        fin_line=[round(v/U_SCALE,U_DEC) for v in fin_eur]
        perf["curve"]["fin_line_js"]=_json.dumps(fin_line); perf["curve"]["has_fin"]=True
        # Échelle conditionnelle : si la variation de période est faible (<2% de la moyenne), une
        # échelle adaptative transforme -0,4% en falaise. On fige alors l'axe Y à moyenne ±2,5%.
        # Dans tous les autres cas (90%), l'échelle adaptative standard reste préférable.
        _vals=[v for v in fin_line if v is not None]
        if len(_vals)>=2:
            _mean=sum(_vals)/len(_vals)
            if _mean and (max(_vals)-min(_vals))/abs(_mean) < 0.02:
                perf["curve"]["y_min"]=round(_mean*0.975,U_DEC)
                perf["curve"]["y_max"]=round(_mean*1.025,U_DEC)
        _psdelta=ps_prorat_total-ps_vm_total
        if ps_list and _psdelta and bool(manifest["reporting"].get("show_ps_corrige", False)):
            _du=_psdelta/U_SCALE
            perf["curve"]["ps_corr_line_js"]=_json.dumps([round(v+_du,U_DEC) for v in fin_line])
            perf["curve"]["has_ps_corr"]=True
        else:
            perf["curve"]["has_ps_corr"]=False
        base_fin=fin_eur[0]; nbf=len(fin_eur)
        _show_bench = bool(manifest["reporting"].get("show_benchmark", False))
        if bperf is not None and base_fin and _show_bench:
            def _fc(i):
                if nbf<=1: return 1.0
                return i/(nbf-1) + 0.10*math.sin(math.pi*i/(nbf-1))   # trajectoire organique, endpoints préservés
            bench=[round(base_fin*(1+(bperf/100.0)*_fc(i))/U_SCALE,U_DEC) for i in range(nbf)]
            perf["curve"]["bench_line_js"]=_json.dumps(bench); perf["curve"]["has_bench"]=True
        else:
            perf["curve"]["has_bench"]=False
    else:
        if perf.get("curve"): perf["curve"]["has_fin"]=False; perf["curve"]["has_bench"]=False; perf["curve"]["has_ps_corr"]=False

    # Widget 3 — commentaire de gestion (slot éditorial P4) + arbitrages de la période
    perf["show_benchmark"] = bool(manifest["reporting"].get("show_benchmark", False))
    perf["commentaire"] = ""   # rédigé en P4 (data-slot éditorial)
    _multi_e = len(manifest["entities"]) > 1
    _arbs = []
    for _e in manifest["entities"]:
        _sxa = suf.get(_e["id"], "")
        for _r in read_sheet(wb, "Arbitrages", _sxa):
            _dc = g(_r, 0); _lbl = s(g(_r, 1)).strip()
            if not _lbl: continue
            _dt = _dc.strftime("%d/%m/%Y") if hasattr(_dc, "strftime") else s(_dc).strip()
            _arbs.append({"date": _dt, "label": _lbl, "entity": (_e["label"] if _multi_e else "")})
    perf["arbitrages"] = _arbs

    # Widget 4 — tableau exhaustif coté par contrat + accordéon lignes (onglet « Lignes — [entité] »)
    from collections import OrderedDict as _OD3
    _lignes_raw = {}
    for _e in manifest["entities"]:
        _rows = []
        for _r in read_sheet(wb, "Lignes", suf.get(_e["id"], "")):
            _ck = _norm(g(_r, 0)); _lbl = s(g(_r, 1)).strip()
            if not _ck or not _lbl: continue
            _rows.append({"ck": _ck, "pk": _norm(s(g(_r, 5)).strip()), "pk_raw": s(g(_r, 5)).strip(),
                "libelle": _lbl, "isin": s(g(_r, 2)).strip() or "&mdash;",
                "_vraw": num(g(_r, 3)) or 0.0, "_praw": num(g(_r, 4)),
                "cls": s(g(_r, 6)).strip()})
        _lignes_raw[_e["id"]] = _rows
    _contracts = _OD3()
    for _eid, _r in pos_raw:
        _key = (_eid, s(g(_r,0)), s(g(_r,1)), s(g(_r,2)))
        _contracts.setdefault(_key, []).append(_r)
    _ent_c = _OD3()
    import datetime as _dtc
    _repd = _dtc.date(*(int(x) for x in manifest["reporting"]["date_reporting"].split("-")))
    def _cmetrics(rows, dated_flows=None):
        nom=sum(num(g(r,10)) or 0 for r in rows); val=sum(num(g(r,12)) or 0 for r in rows)
        vers=sum(num(g(r,17)) or 0 for r in rows); rach=sum(num(g(r,18)) or 0 for r in rows); frais=sum(num(g(r,19)) or 0 for r in rows)
        capital=nom+vers-rach-frais                      # capital investi NET (retraits + frais d'entrée)
        gain=val-capital
        dts=[(pdate(g(r,8)), s(g(r,8)).strip()) for r in rows if pdate(g(r,8))]
        dinv=dmy(min(dts)[0]) if dts else (s(g(rows[0],8)).strip() or "&mdash;")
        # --- Modified Dietz : dénominateur = nominal (au jour d'origine) + Σ flux·poids temporel ---
        denom=capital; days=None
        if dts:
            d0=_dtc.date(*min(dts)[0]); days=(_repd-d0).days
            if dated_flows and days and days>0:
                denom=nom - frais
                for (_fd,_fa) in dated_flows:
                    _t=max(0,(_dtc.date(*_fd)-d0).days)
                    denom += _fa*max(0.0,(days-_t)/days)
        gp=(gain/denom*100) if denom else None
        # --- Perf % YTD (Dietz) : depuis la Valeur 01/01, flux de l'année pondérés ---
        v01s=sum(num(g(r,11)) or 0 for r in rows if num(g(r,11)) is not None)
        ytd=None
        if v01s:
            _ry=_repd.year; _y0=_dtc.date(_ry,1,1); _Ty=(_repd-_y0).days or 1
            _cfy=0.0; _wfy=0.0
            for (_fd,_fa) in (dated_flows or []):
                if _fd[0]==_ry:
                    _ty=max(0,(_dtc.date(*_fd)-_y0).days)
                    _cfy+=_fa; _wfy+=_fa*max(0.0,(_Ty-_ty)/_Ty)
            _dy=v01s+_wfy
            if _dy: ytd=(val-v01s-_cfy)/_dy*100
            ytd_eur=val-v01s-_cfy
        else:
            ytd_eur=None
        gcls="v-pos" if gain>0 else ("v-neg" if gain<0 else "")
        ycls=("v-pos" if ytd_eur and ytd_eur>0 else ("v-neg" if ytd_eur and ytd_eur<0 else "")) if ytd_eur is not None else ""
        return {"nominal":eur(capital),"valeur":eur(val),
                "vers":(eur(vers) if vers else "&mdash;"),"rach":(eur(rach) if rach else "&mdash;"),
                "gain":(f'<span class="{gcls}">{_signed_eur(eur(gain))}</span>' if gcls else eur(gain)),
                "gain_ytd":((f'<span class="{ycls}">{_signed_eur(eur(ytd_eur))}</span>' if ycls else eur(ytd_eur)) if ytd_eur is not None else "&mdash;"),
                "gpct":badge(gp),"gpct_ytd":(badge(ytd) if ytd is not None else "&mdash;"),
                "gpct_ann":"&mdash;",
                "date_inv":dinv,"nant":"&mdash;"}
    def _fmt_lines(rows, base):
        out=[]
        for L in rows:
            vr=L["_vraw"]; pr=L["_praw"]
            pe=(vr - vr/(1+pr/100.0)) if (pr is not None and (1+pr/100.0)) else None
            al=(vr/base*100) if base else None
            out.append({"libelle":L["libelle"],"isin":L["isin"],"valeur":eur(vr),
                "alloc":(f"{al:.1f}%".replace(".",",") if al is not None else "&mdash;"),
                "perfe":(_signed_eur(eur(pe)) if pe else eur(0)),"perf":badge(pr)})
        return out
    _dispo=[]; _mp=[]; _dispo_tot=0.0; _mp_tot=0.0   # demandes client : disponibilités par contrat + détail matières premières
    _entlbl={_e["id"]: _e["label"] for _e in manifest["entities"]}
    for (_eid,_nat,_ass,_int), _grp in _contracts.items():
        _label = f"{NATURE.get(_nat.lower(), _nat)} \u00b7 {_ass}"
        _dfl=[]
        for _ckd in (_norm(_label.replace(" \u00b7 "," — ")), _norm(_label)):
            _dfl += _mvts_dated.get(_eid,{}).get(_ckd,[])
        _cm=_cmetrics(_grp, dated_flows=_dfl); _val=sum(num(g(r,12)) or 0 for r in _grp)
        _ckset={_norm(_label.replace(" \u00b7 "," — ")), _norm(_label)}
        _clines=[L for L in _lignes_raw.get(_eid,[]) if L["ck"] in _ckset]
        from collections import OrderedDict as _ODp
        _pgroups=_ODp()
        for r in _grp:
            _pgroups.setdefault(s(g(r,16)).strip(), []).append(r)   # regroupe les lignes Fin coté par libellé de poche
        _distinct=[p for p in _pgroups if p]
        _sub=None; _poches=[]; _lines=[]
        _pkg=_ODp()
        for L in _clines:
            if L["pk"]: _pkg.setdefault(L["pk"], []).append(L)
        if not _distinct and len(_pkg)>=2:
            # poches déclarées au niveau des LIGNES (contrat mono-ligne en Fin coté) : valeur + actifs par poche
            for _pkn,_pls in _pkg.items():
                _pval=sum(L["_vraw"] for L in _pls)
                _poches.append({"gestion":(_pls[0]["pk_raw"] or _pkn),"sub":None,
                                "date_inv":"&mdash;","nant":"&mdash;","nominal":"&mdash;",
                                "vers":"&mdash;","rach":"&mdash;","valeur":eur(_pval),
                                "gain":"&mdash;","gain_ytd":"&mdash;","gpct":"&mdash;","gpct_ytd":"&mdash;","gpct_ann":"&mdash;",
                                "lines":_fmt_lines(_pls, _pval)})
            _unp=[L for L in _clines if not L["pk"]]
            if _unp: _lines=_fmt_lines(_unp, _val)
            _distinct=[p["gestion"] for p in _poches]        # active le rendu multi-poche
        elif not _distinct:
            _r0=_grp[0]
            _sub=" \u00b7 ".join([x for x in [s(g(_r0,4)).strip(), s(g(_r0,7)).strip()] if x]) or None
            _lines=_fmt_lines(_clines, _val)                 # mono : contrat = poche -> actifs sous le contrat
        else:
            _matched=set()
            for _plabel,_rows in _pgroups.items():
                if not _plabel: continue
                pm=_cmetrics(_rows)
                pm["gestion"]=_plabel or s(g(_rows[0],6)).strip()
                pm["sub"]=" \u00b7 ".join([x for x in [s(g(_rows[0],5)).strip(), s(g(_rows[0],6)).strip(), s(g(_rows[0],7)).strip()] if x])
                _pk=_norm(_plabel)
                _pval=sum(num(g(r,12)) or 0 for r in _rows)
                _pl=[L for L in _clines if L["pk"] and L["pk"]==_pk]
                for L in _pl: _matched.add(id(L))
                pm["lines"]=_fmt_lines(_pl, _pval)
                _poches.append(pm)
            _un=[L for L in _clines if id(L) not in _matched]
            if _un: _lines=_fmt_lines(_un, _val)
        _npoches=(len(_distinct) if _distinct else 1)
        # « Nombre de poches » saisi (col 21) : repli d'affichage si pas de lignes de poche
        if not _distinct:
            _np_in=num(g(_grp[0],21))
            if _np_in and _np_in>1: _npoches=int(_np_in)
        _psd=0.0
        for _p in ps_list:
            if _p.get("eid")!=_eid: continue
            _env=_norm(s(_p.get("env","")))
            if _env and _norm(_ass) and _norm(_ass) in _env:
                _psd += (_p.get("val_prorat",0) or 0) - (_p.get("vm",0) or 0)
        # « Valeur projetée » saisie (col 20) : prime sur le calcul PS si renseignée
        _vp_in=sum(num(g(r,20)) or 0 for r in _grp)
        _c={"label":_label,"interm":_int,"sub":_sub,"npoches":_npoches,"poches":_poches,"lines":_lines,
            "valeur_proj": (eur(_vp_in) if _vp_in else (eur(_val+_psd) if _psd else eur(_val)))}
        _c.update(_cm)
        _ent_c.setdefault(_eid, []).append(_c)
        _dv=sum(L["_vraw"] for L in _clines
                if (L.get("cls")=="Monétaire") or L["libelle"].lower().startswith(("liquid","compte courant","espèces","especes","cash")))
        if _dv:
            _dispo.append({"contrat":_label,"entity":(_entlbl.get(_eid,"") if len(manifest["entities"])>1 else ""),
                           "montant":eur(_dv),"part":(f"{_dv/_val*100:.1f}%".replace(".",",") if _val else "&mdash;")})
            _dispo_tot+=_dv
        for L in _clines:
            if L.get("cls")=="Matières premières":
                _mp.append({"libelle":L["libelle"],"isin":L["isin"],"contrat":_label,
                            "valeur":eur(L["_vraw"]),"perf":badge(L["_praw"])})
                _mp_tot+=L["_vraw"]
    _multi_c = len(manifest["entities"]) > 1
    _cote_table = []
    for _e in manifest["entities"]:
        _cs = _ent_c.get(_e["id"])
        if _cs: _cote_table.append({"entity": (_e["label"] if _multi_c else ""), "contracts": _cs})
    perf["cote_table"] = _cote_table
    perf["dispo"]=({"rows":_dispo,"total":eur(_dispo_tot),
                    "part":(f"{_dispo_tot/g_val*100:.1f}%".replace(".",",") if g_val else "&mdash;")} if _dispo else None)
    perf["mp"]=({"rows":_mp,"total":eur(_mp_tot)} if _mp else None)

    # ---- Bloc PERFORMANCE NON COTÉ : données réelles (onglets Non coté + NC Flux) ----
    # Axe trimestriel figé : démarre à T4 2025, jusqu'au trimestre de reporting.
    # Vues : portefeuille consolidé + une vue par fonds (bascule côté client, cf. performance_nc.html.j2).
    ps = manifest["reporting"]["period_short"]            # ex. "T2-26"
    m = re.match(r"[TS]?(\d)-(\d{2})", ps)
    cur_q = int(m.group(1)) if m else 2
    cur_y = 2000 + int(m.group(2)) if m else 2026
    qs=[]; q,y = 4,2025                                   # départ figé : T4 2025
    while (y,q) <= (cur_y,cur_q):
        qs.append((q,y)); q += 1
        if q > 4: q,y = 1,y+1
    labels=[f"T{q}-{str(y)[2:]}" for (q,y) in qs]
    _QEND={1:(3,31),2:(6,30),3:(9,30),4:(12,31)}
    qends=[(y,_QEND[q][0],_QEND[q][1]) for (q,y) in qs]
    _ry,_rm,_rd = (int(x) for x in manifest["reporting"]["date_reporting"].split("-"))
    rdate=(_ry,_rm,_rd)

    def _nav_at(f, flux, until):
        """NAV d'un fonds à une date : dernière Valorisation <= date (palier) ;
        à défaut cumul des appels <= date (proxy au coût)."""
        valos=[x for x in flux if x["t"]=="Valorisation" and x["d"]<=until]
        if valos: return max(valos, key=lambda x:x["d"])["m"]
        calls=sum(x["m"] for x in flux if x["t"]=="Appel" and x["d"]<=until)
        return calls if calls>0 else None

    def _mk_view(vid, label, series_eur, pe=None):
        """KPIs bandeau (début · MOIC · TVPI · fin base TVPI) + clés période (supervision) + barres."""
        vals=[(None if v is None else float(v)) for v in series_eur]
        nn=[v for v in vals if v is not None] or [0.0]
        scale,unit,dec = (1e6,"M€",2) if max(nn)>=1e6 else (1e3,"K€",0)
        vm=[(None if v is None else round(v/scale,dec)) for v in vals]
        start_v=next((v for v in vals if v is not None), None)
        end_v=next((v for v in reversed(vals) if v is not None), None)
        p=((end_v-start_v)/start_v*100) if start_v else None
        nnm=[v for v in vm if v is not None] or [0]
        _ymin=round(min(nnm)*0.94,dec); _ymax=max(nnm)*1.08
        _lift=round((_ymax-_ymin)*0.11,4)                 # la ligne flotte au-dessus des barres
        line=[(None if v is None else round(v+_lift,4)) for v in vm]
        # --- multiples MOIC (Excel) + TVPI (calculé) ---
        moic=tvpi=None; moic_eur=tvpi_eur=end_tvpi=None
        if pe:
            _c=pe.get("called"); _d=pe.get("dist") or 0.0; _n=pe.get("nav")
            tvpi=pe.get("tvpi")
            if tvpi is None and _c and _n is not None: tvpi=(_d+_n)/_c
            moic=pe.get("moic");  moic=tvpi if moic is None else moic
            moic_eur=(_c*(moic-1)) if (_c and moic is not None) else None
            tvpi_eur=((_d+(_n or 0))-_c) if (_c and _n is not None) else None
            end_tvpi=(_c*tvpi) if (_c and tvpi is not None) else None
        _mcls=lambda m:("c-pos" if (m is None or m>=1) else "c-neg")
        _mcard=lambda m:("card-pos" if (m is None or m>=1) else "card-neg")
        _p0 = (p is None or round(p,2)==0)   # période nulle = pas d'information -> barre
        kpis={"start_value":eur(start_v),
              # clés période (réutilisées par le bloc supervision 01)
              "pct":("&mdash;" if _p0 else pct2(p)),
              "eur":("&mdash;" if (_p0 or start_v is None or end_v is None) else eur(end_v-start_v)),
              "cls":("c-neg" if (p or 0)<0 else "c-pos"),
              "card":("card-neg" if (p or 0)<0 else "card-pos"),
              # bandeau Widget 1
              "moic_mult":(mult(moic) if moic is not None else "&mdash;"),
              "moic_pct":(pct2((moic-1)*100) if moic is not None else "&mdash;"),
              "moic_eur":(_signed_eur(eur(moic_eur)) if moic_eur is not None else "&mdash;"),
              "moic_cls":_mcls(moic),"moic_card":_mcard(moic),
              "tvpi_mult":(mult(tvpi) if tvpi is not None else "&mdash;"),
              "tvpi_pct":(pct2((tvpi-1)*100) if tvpi is not None else "&mdash;"),
              "tvpi_eur":(_signed_eur(eur(tvpi_eur)) if tvpi_eur is not None else "&mdash;"),
              "tvpi_cls":_mcls(tvpi),"tvpi_card":_mcard(tvpi),
              "value":(eur(end_tvpi) if end_tvpi is not None else eur(end_v))}
        return {"id":vid,"label":label,"kpis":kpis,
                "bars":{"labels":labels,"data":vm,"line":line,"unit":unit,"dec":dec,"ymin":_ymin},
                "perf_pct":(None if p is None else round(p,2))}

    # séries € par fonds + stats PE (cumuls, multiples, TRI fourni)
    fund_rows=[]
    for i,f in enumerate(nc_funds):
        flux=sorted(nc_flux.get(f["name"], []), key=lambda x:x["d"])
        serie=[_nav_at(f, flux, qe) for qe in qends]
        if all(v is None for v in serie):
            base=f["vl"] if f["vl"] is not None else f["app"]
            _di=pdate(f.get("date_inv"))
            if _di:
                serie=[(base if qe>=_di else None) for qe in qends]   # pas de flux : palier depuis l'investissement
            else:
                serie=[None]*len(qends)
            if all(v is None for v in serie) and serie: serie[-1]=base   # à minima le trimestre du reporting
        v_r=_nav_at(f, flux, rdate)
        if v_r is not None and serie: serie[-1]=v_r       # dernier trimestre borné à la date de reporting
        called=sum(x["m"] for x in flux if x["t"]=="Appel" and x["d"]<=rdate)
        if called<=0: called=f["app"] or 0.0
        dist=sum(x["m"] for x in flux if x["t"]=="Distribution" and x["d"]<=rdate)
        nav=next((v for v in reversed(serie) if v is not None), None)
        if nav is None: nav=f["vl"] if f["vl"] is not None else called
        eng=f["eng"]
        mill=next((x["d"][0] for x in flux if x["t"]=="Appel"), None)
        _tvpi=(((dist+nav)/called) if called else None)
        _moic=f.get("moic_x") if f.get("moic_x") is not None else _tvpi   # MOIC = multiple BRUT (fonds) ; défaut = net
        _typ=f.get("typ","")
        _direct=(_typ.startswith("titre") if _typ else (f.get("cls")=="Actions non cotées"))   # Titre = détention directe/support ; Fonds = véhicule à engagement
        if _direct: _moic=_tvpi=None
        fund_rows.append({"f":f,"flux":flux,"serie":serie,"called":called,"dist":dist,"nav":nav,
                          "eng":eng,"mill":mill,"moic":_moic,"cible":(None if _direct else f.get("cible_x")),"direct":_direct,
                          "na_reel":f.get("na_reel"),"na_est":f.get("na_est"),
                          "tvpi":_tvpi,
                          "dpi":(None if _direct else ((dist/called) if called else None)),
                          "rvpi":(None if _direct else ((nav/called) if called else None))})

    # série portefeuille = somme des fonds (None si aucun fonds valorisé sur le trimestre)
    port_serie=[]
    for k in range(len(qends)):
        vs=[fr["serie"][k] for fr in fund_rows if fr["serie"][k] is not None]
        port_serie.append(sum(vs) if vs else None)
    _fonds=[fr for fr in fund_rows if not fr["direct"]]   # multiples : fonds seuls (les titres au nominal écrasent vers 1x)
    p_called=sum(fr["called"] for fr in _fonds if fr["called"]) or 0.0
    p_dist=sum(fr["dist"] for fr in _fonds if fr["dist"]) or 0.0
    p_nav=sum(fr["nav"] for fr in _fonds if fr["nav"]) or 0.0
    p_tvpi=((p_dist+p_nav)/p_called) if p_called else None
    _mw=[(fr["moic"],fr["called"]) for fr in _fonds if fr["moic"] is not None and fr["called"]]
    p_moic=(sum(m*c for m,c in _mw)/sum(c for _,c in _mw)) if _mw else p_tvpi
    port_pe={"called":p_called,"dist":p_dist,"nav":p_nav,"moic":p_moic,"tvpi":p_tvpi}
    views=[_mk_view("port","Portefeuille non coté",port_serie,pe=port_pe)]
    for i,fr in enumerate(fund_rows):
        views.append(_mk_view(f"f{i}", fr["f"]["name"], fr["serie"],
                     pe={"called":fr["called"],"dist":fr["dist"],"nav":fr["nav"],
                         "moic":fr["moic"],"tvpi":fr["tvpi"]}))
    pv=views[0]
    pe_pct=pv["perf_pct"]                                  # réutilisé par le bloc Rendement annuel

    # composition du portefeuille (panneau latéral, remplace le benchmark de pairs)
    tot_nav=sum(fr["nav"] for fr in fund_rows if fr["nav"]) or 0.0
    _cib=[(fr.get("cible"),fr["called"]) for fr in fund_rows if fr.get("cible") and fr["called"]]
    p_cible=(sum(c*w for c,w in _cib)/sum(w for _,w in _cib)) if _cib else None
    _allm=[x for fr in fund_rows for x in (fr["moic"],fr["tvpi"],fr.get("cible")) if x is not None]
    _allm+=[x for x in (p_moic,p_tvpi,p_cible) if x is not None]
    _mmax=max(_allm) if _allm else 1.0
    _bw=lambda m:(round(min(100,(m/_mmax)*100)) if (m and _mmax) else 0)
    def _crow(view,name,is_port,weight,moic,tvpi,cible):
        return {"view":view,"name":name,"is_port":is_port,"weight":weight,
                "moic":(mult(moic) if moic is not None else "&mdash;"),
                "tvpi":(mult(tvpi) if tvpi is not None else "&mdash;"),
                "cible":(mult(cible) if cible is not None else None),
                "moic_w":_bw(moic),"tvpi_w":_bw(tvpi),"cible_w":(_bw(cible) if cible else None)}
    compo=[_crow("port","Portefeuille non coté",True,None,p_moic,p_tvpi,p_cible)]
    for i,fr in enumerate(fund_rows):
        compo.append(_crow(f"f{i}",fr["f"]["name"],False,
                     (round(fr["nav"]/tot_nav*100) if tot_nav else None),
                     fr["moic"],fr["tvpi"],fr.get("cible")))

    # page « Détails » — vue portefeuille : synthèse une ligne par fonds
    _sum=lambda k:(sum(fr[k] for fr in fund_rows if fr[k] is not None) if any(fr[k] is not None for fr in fund_rows) else None)
    def _sumf(k):
        _v=[fr[k] for fr in fund_rows if not fr["direct"] and fr[k] is not None]
        return sum(_v) if _v else None
    t_eng=_sumf("eng"); t_called=_sumf("called"); t_dist=_sumf("dist"); t_nav=_sumf("nav")
    t_reste=(t_eng-t_called) if (t_eng is not None and t_called is not None) else None
    _tris=[(fr["f"]["tri"],fr["called"]) for fr in fund_rows if fr["f"]["tri"] is not None and fr["called"]]
    t_tri=(sum(t*c for t,c in _tris)/sum(c for _,c in _tris)) if _tris else None
    # unité compacte unique pour tout le tableau (lisibilité : pas de montants € pleins sur 10 colonnes)
    _vals=[abs(x) for fr in fund_rows for x in (fr["eng"],fr["called"],fr["dist"],fr["nav"]) if x]
    _sc,_un,_dc=(1e6,"M€",2) if (_vals and max(_vals)>=1e6) else (1e3,"K€",0)
    def eur_u(v):
        if v is None: return "&mdash;"
        n=round(v/_sc,_dc)
        if n==0: return "&mdash;"
        sg="−" if n<0 else ""
        return f"{sg}{abs(n):.{_dc}f}".replace(".",",")+f" {_un}"
    summary={"rows":[],"total":None,"unit":_un}
    for fr in fund_rows:
        reste=(fr["eng"]-fr["called"]) if fr["eng"] is not None else None
        summary["rows"].append({"name":fr["f"]["name"],"gest":fr["f"]["gest"],
            "mill":(str(fr["mill"]) if fr["mill"] else "&mdash;"),
            "eng":eur_u(fr["eng"]),"called":eur_u(fr["called"] or None),"reste":eur_u(reste),
            "dist":(eur_u(fr["dist"]) if fr["dist"] else "&mdash;"),"nav":eur_u(fr["nav"]),
            "tvpi":mult(fr["tvpi"]),"dpi":mult(fr["dpi"]),"tri":tri_fmt(fr["f"]["tri"])})
    if fund_rows:
        summary["total"]={"eng":eur_u(t_eng),"called":eur_u(t_called),"reste":eur_u(t_reste),
            "dist":(eur_u(t_dist) if t_dist else "&mdash;"),"nav":eur_u(t_nav),
            "tvpi":mult(((t_dist or 0)+(t_nav or 0))/t_called if t_called else None),
            "dpi":mult((t_dist or 0)/t_called if (t_called and t_dist) else None),
            "tri":(("≈ "+tri_fmt(t_tri)) if t_tri is not None else "&mdash;")}

    # prochaine échéance prévisionnelle — vue portefeuille (tous fonds confondus)
    _fut=[(fr["f"]["name"],x) for fr in fund_rows for x in fr["flux"]
          if x["d"]>rdate and x["t"] in ("Appel prévu","Distribution prévue")]
    if _fut:
        _fn,_x=min(_fut,key=lambda t:t[1]["d"])
        summary["next"]=(f"Prochaine échéance : <strong>{_x['t'].lower()}</strong> de "
                         f"<strong>{eur(_x['m'])}</strong> — {_fn}, le {dmy(_x['d'])}.")
    else:
        summary["next"]=None

    # texte de synthèse commercial (panneau composition), déterministe depuis les données
    pitch=None
    if fund_rows:
        _LBLNC={"Private Equity":"private equity","Dette privée":"dette privée",
                "Immo non coté":"immobilier non coté","Infrastructures":"infrastructures"}
        _cls=[_LBLNC.get(c,c.lower()) for c in ["Private Equity","Dette privée","Immo non coté","Infrastructures"]
              if any(fr["f"]["cls"]==c for fr in fund_rows)]
        _clstxt=(_cls[0] if len(_cls)==1 else " et ".join(_cls) if len(_cls)==2
                 else ", ".join(_cls[:-1])+" et "+_cls[-1]) if _cls else "non coté"
        _nf=len(fund_rows)
        pitch=(f"Votre allocation non cotée regroupe <strong>{_nf} fonds</strong>" if _nf>1
               else "Votre allocation non cotée repose sur <strong>1 fonds</strong>")
        pitch+=f" — {_clstxt} — pour un engagement total de <strong>{eur(t_eng)}</strong>"
        if t_eng and t_called is not None:
            pitch+=f", appelé à hauteur de <strong>{t_called/t_eng*100:.0f}%</strong>".replace(".",",")
        pitch+="."
        if t_reste and t_reste>0:
            pitch+=(f" Le solde de <strong>{eur(t_reste)}</strong> sera appelé progressivement selon "
                    "l'échéancier de chaque millésime — un déploiement par paliers qui lisse les points "
                    "d'entrée et construit la performance dans la durée.")
        pitch+=(" Chaque ligne fait l'objet d'un suivi individualisé par nos équipes : sélectionnez un "
                "fonds pour consulter sa valorisation et son échéancier de flux.")

    # page « Détails » — vue fonds : échéancier de flux daté (réalisé + prévisionnel)
    schedules=[]
    for i,fr in enumerate(fund_rows):
        rows=[]; cum_calls=0.0
        for x in fr["flux"]:
            prevu=x["t"].endswith("prévu") or x["t"].endswith("prévue")
            reste_ap=None
            if x["t"] in ("Appel","Appel prévu"):
                cum_calls+=x["m"]
                if fr["eng"] is not None: reste_ap=fr["eng"]-cum_calls
            rows.append({"date":dmy(x["d"]),"type":x["t"],"montant":eur(x["m"]),
                         "prevu":prevu,"future":(x["d"]>rdate),
                         "reste":(eur(reste_ap) if reste_ap is not None else "")})
        _futf=[x for x in fr["flux"] if x["d"]>rdate and x["t"] in ("Appel prévu","Distribution prévue")]
        if _futf:
            _x=min(_futf,key=lambda k:k["d"])
            nxt=(f"Prochaine échéance : <strong>{_x['t'].lower()}</strong> de "
                 f"<strong>{eur(_x['m'])}</strong>, le {dmy(_x['d'])}.")
        elif rows:
            nxt="Aucune échéance prévisionnelle saisie pour ce fonds."
        else:
            nxt=None
        schedules.append({"view":f"f{i}","name":fr["f"]["name"],"rows":rows,"next":nxt})

    # Widget 4 — détail par contrat (accordéon) : ligne synthèse + dépliant échéancier/stats
    _by_eid={}
    for i,fr in enumerate(fund_rows):
        _nar=fr["na_reel"]
        if _nar is None and fr["eng"] is not None and fr["called"] is not None:
            _nar=fr["eng"]-fr["called"]                     # non appelé réel = engagé − appelé
        _gn=fr["f"]["gest"]; _nm=fr["f"]["name"]
        _mere=(f"{_gn} - {_nm}" if _gn else _nm)
        _seg=fr["f"].get("segment",""); _dur=fr["f"].get("duree","")
        # Nomenclature courte : les valeurs cible/durée arrivent pré-formatées de l'Excel ("1,8x", "10y") — aucun formatage à recalculer
        _sub=" · ".join([x for x in [_seg, fr["f"]["strat"],
                         (("Cible "+fr["f"]["cible"]) if fr["f"]["cible"] else ""),
                         _dur] if x])
        _dinv_nc = fr["f"].get("date_inv","") or (str(fr["mill"]) if fr.get("mill") else "&mdash;")
        _pdn = pdate(_dinv_nc)
        if _pdn: _dinv_nc = dmy(_pdn)
        _direct=fr.get("direct")
        _cible_disp=("&mdash;")
        if _direct:
            if fr["f"]["tri"] is not None: _cible_disp=f'TRI {tri_fmt(fr["f"]["tri"])}'
        elif fr.get("cible") is not None: _cible_disp=mult(fr["cible"])
        elif fr["f"]["cible"]: _cible_disp=fr["f"]["cible"]
        _d={
            "name":_mere,"sub":_sub,"cls":fr["f"]["cls"],"is_titre":_direct,"cible":_cible_disp,
            "date_inv":_dinv_nc,
            "eng":(eur_u(fr["eng"]) if not _direct else "&mdash;"),
            "nominal":(eur_u(fr["eng"]) if _direct else "&mdash;"),
            "called":(eur_u(fr["called"] or None) if not _direct else "&mdash;"),
            "na_reel":(eur_u(_nar) if not _direct else "&mdash;"),
            "na_est":(eur_u(fr["na_est"]) if (fr["na_est"] is not None and not _direct) else "&mdash;"),
            "moic":(mult(fr["moic"]) if fr["moic"] is not None else "&mdash;"),
            "vl":eur_u(fr["nav"]),
            "flux":(schedules[i]["rows"] if not _direct else []),"next":(schedules[i]["next"] if not _direct else None),
            "stats":{"tvpi":mult(fr["tvpi"]),"dpi":mult(fr["dpi"]),
                     "rvpi":mult(fr["rvpi"]),"tri":tri_fmt(fr["f"]["tri"])}}
        _by_eid.setdefault(fr["f"]["eid"], []).append(_d)
    _multi_nc = len(manifest["entities"]) > 1
    nc_detail=[]; nc_detail_t=[]
    for _e in manifest["entities"]:
        if _e["id"] in _by_eid:
            _rf=[d for d in _by_eid[_e["id"]] if not d["is_titre"]]
            _rt=[d for d in _by_eid[_e["id"]] if d["is_titre"]]
            if _rf: nc_detail.append({"entity": (_e["label"] if _multi_nc else ""), "rows": _rf})
            if _rt: nc_detail_t.append({"entity": (_e["label"] if _multi_nc else ""), "rows": _rt})
    _tit=[fr for fr in fund_rows if fr["direct"]]
    nc_titres_total={"nominal":eur_u(sum(fr["eng"] or 0 for fr in _tit)),
                     "vl":eur_u(sum(fr["nav"] or 0 for fr in _tit))} if _tit else None
    _t_nar=(t_eng-t_called) if (t_eng is not None and t_called is not None) else None
    _t_nae=None
    if any(fr["na_est"] is not None for fr in fund_rows):
        _t_nae=sum(fr["na_est"] for fr in fund_rows if fr["na_est"] is not None)
    nc_detail_total={"eng":eur_u(t_eng),"called":eur_u(t_called),
        "na_reel":eur_u(_t_nar),"na_est":(eur_u(_t_nae) if _t_nae is not None else "&mdash;"),
        "moic":(mult(p_moic) if p_moic is not None else "&mdash;"),"vl":eur_u(t_nav)}

    # objet JS embarqué (bascule de vue côté client, aucune dépendance réseau)
    _views_js=[{"id":v["id"],"label":v["label"],"kpis":v["kpis"],"bars":v["bars"]} for v in views]
    _pv_data=pv["bars"]["data"]
    _pv_start_idx=next((i for i,v in enumerate(_pv_data) if v is not None), None)
    _pv_start_label=labels[_pv_start_idx] if _pv_start_idx is not None else labels[-1]
    data["perf_nc"]={
        "kpis":pv["kpis"],
        "bars":{"labels_js":_json.dumps(pv["bars"]["labels"],ensure_ascii=False),
                "data_js":_json.dumps(pv["bars"]["data"]),"line_js":_json.dumps(pv["bars"]["line"]),
                "unit":pv["bars"]["unit"],"dec":pv["bars"]["dec"],"ymin":pv["bars"]["ymin"],
                "start_label":_pv_start_label},
        "views_js":_json.dumps(_views_js,ensure_ascii=False),
        "compo":compo,"summary":summary,"schedules":schedules,"pitch":pitch,
        "detail":nc_detail,"detail_total":nc_detail_total,"detail_titres":nc_detail_t,
        "titres_total":nc_titres_total,"detail_unit":_un,
        "has_funds":bool(fund_rows)}

    # ---- Bloc Rendement annuel (cartes courtes / bar chart + tableau synthétique 3 poches) ----
    if "Historique" in wb.sheetnames:
        GREEN="#2D6E4E"; RED="#8B2E2E"; GOLD="#B8975A"
        # Onglet Historique : Année | Perf coté (%) | Perf non coté (%) [opt] | Commentaire [opt]
        past=[]  # (label, perf_cote%, perf_nc% | None, commentaire)
        for row in wb["Historique"].iter_rows(min_row=4, values_only=True):
            if not row or row[0] in (None,""): continue
            rc=num(row[1])
            if rc is None: continue
            yl=str(int(row[0])) if isinstance(row[0],(int,float)) else str(row[0]).strip()
            c2=row[2] if len(row)>2 else None
            c3=row[3] if len(row)>3 else None
            rnc=num(c2)
            if rnc is not None:
                com=str(c3).strip() if c3 not in (None,"") else ""
            else:
                com=str(c2).strip() if c2 not in (None,"") else ""   # 3e col texte = commentaire
            past.append((yl, round(rc,2), (round(rnc,2) if rnc is not None else None), com))
        cur=manifest["reporting"]["date_reporting"][:4]
        _annual = manifest["reporting"]["date_reporting"][5:7] == "12"
        _rmonth = int(manifest["reporting"]["date_reporting"][5:7])
        _wcur = 1.0 if _annual else round(_rmonth/12.0,4)
        # poids des poches (valeurs courantes) pour le rendement global
        _v_cote=float(cat_brut_global.get("financier_cote",0) or 0)
        _v_nc=float(cat_brut_global.get("non_cote",0) or 0)
        _has_nc = (_v_nc>0) or any(t[2] is not None for t in past)
        _den=_v_cote+_v_nc
        _w_cote = (_v_cote/_den) if _den>0 else 1.0
        _w_nc = (_v_nc/_den) if _den>0 else 0.0
        # série poche cotée (base, toujours présente) : années passées + année courante (YTD)
        labels=[t[0] for t in past]
        cote_v=[t[1] for t in past]; cote_w=[1.0]*len(past)
        if ytd_pct is not None:
            labels=labels+[cur if _annual else f"{cur} (YTD)"]
            cote_v=cote_v+[round(ytd_pct,2)]; cote_w=cote_w+[_wcur]
        # série poche non cotée : suffixe contigu (le plus récent) d'années renseignées + YTD non coté
        nc_v=[]; nc_w=[]
        if _has_nc:
            _nc_cur = round(pe_pct,2) if (_v_nc>0 and pe_pct is not None) else None
            seq=[(t[2],1.0) for t in past]
            if ytd_pct is not None: seq=seq+[(_nc_cur,_wcur)]
            tail=[]
            for val,w in reversed(seq):
                if val is None: break
                tail.append((val,w))
            tail.reverse()
            nc_v=[v for v,_ in tail]; nc_w=[w for _,w in tail]
        if len(cote_v) >= 2:
            colors=[(GOLD if i==len(cote_v)-1 else (GREEN if v>=0 else RED)) for i,v in enumerate(cote_v)]
            FULL=20.0  # ±20% remplit la mini-barre des cartes
            cards=[]
            for i,(yl,v) in enumerate(zip(labels,cote_v)):
                _last=(i==len(cote_v)-1)
                cards.append({"year":yl,"perf":pct2(v),
                    "cls":("c-pos" if v>=0 else "c-neg"),
                    "card_cls":("card-gold" if _last else ("card-pos" if v>=0 else "card-neg")),
                    "bar_cls":("bench-bar-pos" if v>=0 else "bench-bar-neg"),
                    "width":round(min(100,abs(v)/FULL*100))})
            H={"mode":("cards" if len(cote_v)<=HIST_CARDS_MAX else "chart"),"cards":cards,
               "labels_js":_json.dumps(labels,ensure_ascii=False),
               "data_js":_json.dumps(cote_v),"colors_js":_json.dumps(colors)}
            def _vivid(hx):                                   # survol : couleur qui "pop" (plus saturée, pas délavée)
                import colorsys
                hx=hx.lstrip("#"); r,g,b=[int(hx[i:i+2],16)/255.0 for i in (0,2,4)]
                h,l,sat=colorsys.rgb_to_hls(r,g,b)
                sat=min(1.0, sat*1.7+0.12); l=min(0.55, l*1.18)
                r,g,b=colorsys.hls_to_rgb(h,l,sat)
                return "#%02X%02X%02X"%(int(r*255),int(g*255),int(b*255))
            H["colors_hover_js"]=_json.dumps([_vivid(c) for c in colors])
            if H["mode"]=="chart" and past:
                b=max(past,key=lambda t:t[1]); w=min(past,key=lambda t:t[1])
                H["stats"]={"best":{"year":b[0],"perf":pct2(b[1]),"note":b[3]},
                            "worst":{"year":w[0],"perf":pct2(w[1]),"note":w[3]}}
                # --- Tableau synthétique des rendements ANNUALISÉS (poche cotée / non cotée / global) ---
                def _ann(rets, wts, N=None):
                    if not rets: return None
                    rs=[r/100.0 for r in rets]
                    if N is None:                       # depuis lancement : annualisé sur la durée totale
                        acc=1.0; T=0.0
                        for r,w_ in zip(rs,wts): acc*=(1.0+r); T+=w_
                        return (acc**(1.0/T)-1.0) if T>0 else None
                    acc=1.0; rem=float(N)               # période glissante de N années (fraction gérée)
                    for r,w_ in zip(reversed(rs),reversed(wts)):
                        if w_<=0: continue
                        take=min(w_,rem); acc*=(1.0+r)**(take/w_); rem-=take
                        if rem<=1e-9: break
                    if rem>1e-9: return None             # historique insuffisant
                    return acc**(1.0/N)-1.0
                def _pf(x):
                    if x is None: return "—"
                    return (f"{x*100:.2f}").replace("-","−").replace(".",",")
                def _gm(vals):   # moyenne annualisée (géométrique) d'une liste de rendements annuels (%)
                    if not vals: return None
                    acc=1.0
                    for v in vals: acc*=(1.0+v/100.0)
                    return acc**(1.0/len(vals))-1.0
                _past_cote=[t[1] for t in past]; _past_nc=[t[2] for t in past]
                _cur_cote=(ytd_pct/100.0) if ytd_pct is not None else None
                _cur_nc=(pe_pct/100.0) if (_v_nc>0 and pe_pct is not None) else None
                _heads=[]; _cote_by=[]; _nc_by=[]; _glob_by=[]
                # YTD = année civile en cours
                if _cur_cote is not None:
                    nv=_cur_nc if _has_nc else None
                    gv=(_w_cote*_cur_cote+_w_nc*nv) if (_has_nc and nv is not None) else None
                    _heads.append("YTD"); _cote_by.append(_pf(_cur_cote)); _nc_by.append(_pf(nv)); _glob_by.append(_pf(gv))
                # 3/5/10 ans = moyenne annualisée des N dernières années civiles révolues
                for N in (3,5,10):
                    if len(_past_cote)<N: continue
                    cv=_gm(_past_cote[-N:])
                    _ncsl=_past_nc[-N:]
                    nv=_gm(_ncsl) if (_has_nc and all(x is not None for x in _ncsl)) else None
                    gv=(_w_cote*cv+_w_nc*nv) if (_has_nc and nv is not None) else None
                    _heads.append(f"{N} ans"); _cote_by.append(_pf(cv)); _nc_by.append(_pf(nv)); _glob_by.append(_pf(gv))
                _dr=manifest["reporting"]["date_reporting"]
                H["synth"]={"heads":_heads,"cote":_cote_by,"has_nc":_has_nc,"nc":_nc_by,"glob":_glob_by,
                            "wcote":round(_w_cote*100),"wnc":round(_w_nc*100),
                            "caption":("YTD = ann\u00e9e civile en cours (au %s/%s/%s) ; 3/5/10 ans = moyenne annualis\u00e9e (g\u00e9om\u00e9trique) des N derni\u00e8res ann\u00e9es civiles r\u00e9volues. Rendement global = moyenne pond\u00e9r\u00e9e des deux poches." % (_dr[8:10],_dr[5:7],_dr[0:4]))}
                # ligne "parcours" qui suit le sommet des barres (visuel repris du PE), légèrement relevée
                # ligne "parcours" (axe %) : années POSITIVES = valeur réelle (inchangé) ;
                # années NÉGATIVES = point précédent + rendement (repli doux depuis le niveau atteint,
                # ex. 11,2% puis -4,8% -> 6,4%). On marque le repli sans plonger en absolu.
                _line=[]; _prev=None
                for _v in cote_v:
                    _pt=_v if _v>=0 else ((_prev if _prev is not None else 0.0)+_v)
                    _line.append(round(_pt,2)); _prev=_pt
                # on remonte toute la courbe d'un décalage constant -> elle flotte au-dessus des barres
                # (forme + creux inchangés). Ajuster le facteur 0.7 pour monter/descendre l'ensemble.
                _shift=round(max(max(cote_v),1.0)*0.7,2)
                H["line_js"]=_json.dumps([round(x+_shift,2) for x in _line])
                # événements descriptifs : commentaires de l'onglet Historique -> "Année → commentaire"
                _evs=[]
                for _i in range(len(labels)):
                    if _i < len(past) and past[_i][3]:
                        _evs.append({"i":_i,"val":cote_v[_i],"txt":(labels[_i]+" → "+past[_i][3])})
                H["events_js"]=_json.dumps(_evs, ensure_ascii=False)
                # texte de lecture de marché Rhétorès sur la période (carte latérale, façon coté/PE)
                _ann_all=_ann(cote_v,cote_w,None)
                _neg=[labels[i] for i,v in enumerate(cote_v) if v<0]
                _b=H["stats"]["best"]; _w=H["stats"]["worst"]
                _pa=(f"{_ann_all*100:.1f}".replace(".",",")+" %") if _ann_all is not None else "&mdash;"
                _negph=(" Le portefeuille a absorb\u00e9 %d ann\u00e9e(s) de repli (dont %s \u00e0 %s)," % (len(_neg), _w["year"], _w["perf"])) if _neg else " Le portefeuille n'a connu aucune ann\u00e9e n\u00e9gative sur la p\u00e9riode,"
                H["nyears"]=len(labels)
                H["market_text"]=("Depuis l'entr\u00e9e en relation (%s\u2013%s), les march\u00e9s ont altern\u00e9 phases de hausse soutenue et corrections marqu\u00e9es.%s "
                    "tout en d\u00e9livrant un rendement annualis\u00e9 de <strong>%s par an</strong> ; la meilleure ann\u00e9e reste %s (%s). "
                    "Cette trajectoire illustre l'approche de long terme de Rh\u00e9tor\u00e8s : capter la performance des march\u00e9s "
                    "tout en amortissant les chocs par une allocation diversifi\u00e9e." % (past[0][0], cur, _negph, _pa, _b["year"], _b["perf"]))
            # Bloc Rendement annuel SUPPRIMÉ (déprécié au profit du tableau Historique du
            # patrimoine, qui porte déjà la performance annuelle et huit colonnes de plus).
            # Seule la SYNTHÈSE ANNUALISÉE est conservée : elle n'a aucun équivalent ailleurs.
            # Elle est relogée en widget du bloc supervision, SOUS le tableau Historique.
            # Le reste du calcul de H (graphe, repères, texte de marché) demeure mais n'est plus
            # consommé — laissé en place à dessein : le retirer serait un risque sans gain.
            data["annualise"]=H.get("synth")

    # ---- Auto-contrôle comptable (identités) ----
    def _chk(ok, lbl, detail=""):
        qc.append((ok, lbl, detail)); return ok
    qc=[]; TOL=1.5
    net_total = tot_assets - tot_debt
    SB=sum(cat_brut_global.values()); _chk(abs(SB-tot_assets)<=TOL, "Σ brut catégories = actif brut", f"{SB:.0f} vs {tot_assets:.0f}")
    D=sum(debt_by_cat_global.values()); _chk(abs(D-tot_debt)<=TOL, "Σ dettes réparties = dette totale", f"{D:.0f} vs {tot_debt:.0f}")
    Ne=sum(cr["total"]["net"] for cr in cards_raw.values()); _chk(abs(Ne-net_total)<=TOL, "Σ nets entités = actif net", f"{Ne:.0f} vs {net_total:.0f}")
    F=sum(agg_cls.values()); FC=cat_brut_global["financier_cote"]+cat_brut_global["non_cote"]
    _chk(abs(F-FC)<=TOL, "donut classes = financier (coté+non coté)", f"{F:.0f} vs {FC:.0f}")
    ce=perf.get("curve_end_eur")
    if ce is not None: _chk(abs(ce-net_total)<=max(1.0,0.01*abs(net_total)), "courbe (dernier point) ≈ actif net", f"{ce:.0f} vs {net_total:.0f}")
    for cid,dn in data["donuts"].items():
        ssum=sum(L["pct"] for L in dn["legend"]); _chk(98<=ssum<=102, f"donut {cid} ≈ 100%", f"{ssum}%")
    n_ko=sum(1 for ok,_,_ in qc if not ok)
    print(f"── Contrôles comptables : {len(qc)-n_ko}/{len(qc)} OK ──")
    for ok,lbl,det in qc:
        if not ok: print(f"  ✗ {lbl} : {det}")
    if n_ko==0: print("  ✓ Toutes les identités sont vérifiées.")

    # --- QC éditorial : perf ligne manquante (hors liquidités) — avertissement, pas un contrôle comptable ---
    _missing_perf = 0
    for _eid, _rows in _lignes_raw.items():
        for L in _rows:
            _lbl = (L.get("libelle") or "").strip().lower()
            if _lbl.startswith(("liquid", "compte courant", "espèces", "especes", "cash")):
                continue
            if L.get("_praw") is None:
                _missing_perf += 1
    if _missing_perf:
        print(f"  ⚠ {_missing_perf} ligne(s) de détail sans performance (hors liquidités) "
              f"— vérifier la colonne « Perf % » de l\'onglet Lignes avant livraison.")
    if manifest["reporting"].get("mode","presentee")=="presentee" and data.get("perf") and not data["perf"].get("commentaire"):
        print("  ⚠ Commentaire de gestion absent (placeholder P4 encore visible) — à rédiger avant livraison en présentée.")

    qc_comment = "<!-- QC: " + (f"{n_ko} écart(s) — " + "; ".join(f"{lbl} ({det})" for ok,lbl,det in qc if not ok) if n_ko else "toutes identités OK") + " -->\n"

    # ---- Comparaison N-1 : snapshots par client/période ----
    import os as _os, glob as _glob
    pp_id = next((e["id"] for e in manifest["entities"] if e["type"]=="pp"), manifest["entities"][0]["id"])
    snapdir = "snapshots"; cur_date = manifest["reporting"]["date_reporting"]
    cur_w = {c:(cat_brut_global[c]/tot_assets*100 if tot_assets else 0) for c in A.CATEGORY_ORDER if c!="dettes"}
    snap = {"period_short":manifest["reporting"]["period_short"],"period_long":manifest["reporting"]["period_long"],
            "date_reporting":cur_date,"actif_brut":tot_assets,"actif_net":net_total,"dettes":tot_debt,"cat_weights":cur_w,"ps_status":ps_status_cur}
    prev=None
    for f in _glob.glob(_os.path.join(snapdir, f"{pp_id}_*.json")):
        try: sj=_json.load(open(f,encoding="utf-8"))
        except Exception: continue
        if sj.get("date_reporting","") < cur_date and (prev is None or sj["date_reporting"]>prev["date_reporting"]): prev=sj
    if prev:
        pn=prev.get("actif_net") or 0; dnet=net_total-pn
        raw=[]
        for c in A.CATEGORY_ORDER:
            if c=="dettes": continue
            d=cur_w.get(c,0)-prev.get("cat_weights",{}).get(c,0)
            if abs(d)>=0.5: raw.append((c,d))
        raw.sort(key=lambda t:abs(t[1]), reverse=True)
        cats=[{"label":A.CATEGORY_LABELS[c],"delta":("+" if d>=0 else "−")+f"{abs(d):.1f}".replace(".",",")+" pts",
               "cls":"c-pos" if d>=0 else "c-neg","arrow":("▲" if d>=0 else "▼")} for c,d in raw[:4]]
        data["compare_n1"]={"prev_period":prev.get("period_long",prev.get("period_short")),
            "net_eur":("+" if dnet>=0 else "−")+eur(abs(dnet)),"net_cls":"c-pos" if dnet>=0 else "c-neg",
            "net_pct":(pct2(dnet/pn*100) if pn else "&mdash;"),"cats":cats}
    try:
        _os.makedirs(snapdir, exist_ok=True)
        _json.dump(snap, open(_os.path.join(snapdir,f"{pp_id}_{snap['period_short']}.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=2)
    except Exception: pass

    # rendu via la banque (structure identique à P1)
    env = Environment(loader=FileSystemLoader(str(A.BANK)), undefined=StrictUndefined,
                      autoescape=False, trim_blocks=True, lstrip_blocks=True, keep_trailing_newline=True)
    env.filters["frdate"]=A.frdate
    rep=dict(manifest["reporting"]); rep["client_label_html"]=A.client_label_html(rep["client_label"])
    # ---- Bloc « Patrimoine sous supervision » (version présentée) : KPI + donuts 2-4 (€) ----
    _ck = data["perf"].get("kpis", {}) if data.get("perf") else {}
    _nk = data["perf_nc"]["kpis"] if data.get("perf_nc") else None
    sup_kpis = {
        "value": eur(tot_assets),
        "cote_pct": _ck.get("ytd_pct","&mdash;"), "cote_eur": _signed_eur(_ck.get("ytd_eur","&mdash;")),
        "cote_cls": _ck.get("ytd_cls",""), "cote_card": _ck.get("ytd_card","card-pos"),
        "cote_label": _ck.get("ytd_label_cote","Performance YTD — coté"),
        "nc_pct": (_nk.get("pct","&mdash;") if _nk else "&mdash;"),
        "nc_eur": (_signed_eur(_nk.get("eur","&mdash;")) if _nk else "&mdash;"),
        "nc_cls": (_nk.get("cls","") if _nk else ""),
        "nc_card": (_nk.get("card","card-pos") if _nk else "card-pos"),
        # "profil" (SRI moyen) : ajouté au Lot C
    }
    # SRI moyen pondéré (patrimoine) -> profil de risque
    if sri_den > 0:
        import math as _math
        _sri_avg = int(_math.ceil(sri_num/sri_den)); _sri_avg = max(1, min(7, _sri_avg))
        sup_kpis["profil"] = sri_label(_sri_avg)
        sup_kpis["sri"] = _sri_avg
    # Donut 1 — double donut : classes en € (Court terme = liquidités + monétaire) + arcs Coté/Non coté
    _inner = dict(agg_cls)
    _ct = _inner.pop("Monétaire", 0) + float(cat_brut_global.get("liquidites", 0) or 0)
    if _ct > 0: _inner["Court terme"] = _inner.get("Court terme", 0) + _ct
    _nc_names = set(agg_nc_cls.keys())
    _double = make_double_donut(_inner, _nc_names)
    # Donuts 2-4 (avec colonne €)
    sup_donuts = {}
    for _cid, _agg, _cm in [("sup_geo",agg_geo,GEO_COLORS),("sup_env",agg_env,ENV_COLORS),("sup_part",agg_part,DEPO_COLORS)]:
        _dd = make_donut(_agg, _cm)
        if _dd: sup_donuts[_cid]=_dd
    data["supervision"]={"kpis":sup_kpis,"double_donut":_double,"donuts":sup_donuts,
                         "histo_pat":data.get("histo_pat") or [],
                         "annualise":data.get("annualise")}

    layout=A.compute_layout(manifest["blocs_enabled"])
    bloc_num={b["name"]:b["display_num"] for b in layout if b["display_num"] is not None}
    ctx={"reporting":rep,"blocs_enabled":manifest["blocs_enabled"],"bloc_num":bloc_num,
         "entities":manifest["entities"],"category_order":A.CATEGORY_ORDER,
         "category_labels":A.CATEGORY_LABELS,"colspans":A.COLSPANS,"data":data}
    html=env.get_template("base.html.j2").render(**ctx)
    open(outp,"w",encoding="utf-8").write(qc_comment + html)
    print(f"OK reporting rempli: {outp} ({len(html)} octets)")
    print(f"   actif brut {eur(tot_assets)} · dettes {eur(-tot_debt)} · net {eur(tot_assets-tot_debt)}")

if __name__=="__main__": main()
